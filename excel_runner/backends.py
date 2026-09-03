"""Backend primitives: openpyxl (file) and xlwings (live Excel). See Spec sec 3.

Plain functions, not classes, one per primitive operation, so swapping or testing either side
never requires touching action code (PRD sec 6.1). Three naming tiers, not two — "COM" alone
would be inaccurate, since there is no COM on macOS at all, only Apple Events; xlwings abstracts
that difference, which is exactly why it was chosen over raw `win32com` (PRD sec 4):
- File-backend functions (openpyxl) are unprefixed.
- `xlw_`-prefixed functions use xlwings' portable, cross-platform API — the normal live-Excel
  case, works identically on Windows and macOS.
- `com_`-prefixed functions are the genuine exception: they reach through xlwings' `.api`
  escape hatch for something only the raw, Windows-only COM object can do (PRD sec 7's
  `recalculate` full/full_rebuild modes are the known example) — not the default, added only
  where actually needed.
"""

import logging
import re
import shutil
import time
from typing import Any, Literal

import openpyxl
import xlwings as xw
from openpyxl.cell.cell import Cell
from openpyxl.cell.read_only import ReadOnlyCell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)

_SingleCell = (Cell, ReadOnlyCell)
_WHOLE_COLUMN_RE = re.compile(r"^([A-Za-z]+):([A-Za-z]+)$")
_WHOLE_ROW_RE = re.compile(r"^(\d+):(\d+)$")


def open_workbook(
    path: str, mode: Literal["read_only", "read_write"], data_only: bool = True
) -> Workbook:
    """Open an existing workbook file.

    Creating a workbook that doesn't exist yet is a `workbooks:` registry concern
    (`create_if_missing`), handled by session management once it exists — not this function.

    `data_only` — not `read_only` — is the sole determinant of whether a formula cell's
    `.value` reads back as the formula text or its cached computed value (confirmed via
    runtime probing across all `read_only`/`data_only` combinations; `read_only` has no
    effect on this at all). Defaults to True (values), matching every other action's normal
    expectation of reading a computed result, not formula source — `read_range` and
    `read_metadata(target: cells)` are the only two actions that ever need the opposite, via
    their own `formula: true` param (`open_workbook_for_formula_read`), not this default.

    Args:
        path: Path to an existing workbook file.
        mode: "read_only" opens without allowing writes; "read_write" allows them.
        data_only: True (default) reads formula cells as their cached computed value; False
            reads the formula text instead.

    Returns:
        The opened openpyxl Workbook.

    Raises:
        FileNotFoundError: If path does not exist.
    """
    return openpyxl.load_workbook(
        path, read_only=(mode == "read_only"), data_only=data_only
    )


def open_workbook_for_formula_read(path: str) -> Workbook:
    """Open a fresh, read-only, `data_only=False` view of a workbook file, purely to read
    formula text for a `formula: true` request — never the session's main handle, which stays
    on its own `data_only` setting for the rest of the run (data_only is a load-time decision,
    not something togglable on an already-open handle).

    Args:
        path: Path to an existing workbook file — the session's current (possibly scratch)
            path, saved first if the session has unsaved writes.

    Returns:
        A throwaway openpyxl Workbook, read-only and data_only=False. Caller is responsible
        for closing it once done reading.
    """
    return openpyxl.load_workbook(path, read_only=True, data_only=False)


def create_workbook(path: str, template_path: str | None = None) -> None:
    """Create a new workbook file — blank, or copied from a template.

    Args:
        path: Where to create the new workbook.
        template_path: If given, copy this existing workbook's content instead of a blank one.
    """
    if template_path is not None:
        shutil.copy2(template_path, path)
    else:
        openpyxl.Workbook().save(path)


def save_workbook(workbook: Workbook, path: str) -> None:
    """Save a workbook to the given path.

    Args:
        workbook: The workbook to save.
        path: Destination path.
    """
    workbook.save(path)


def close_workbook(workbook: Workbook) -> None:
    """Close a workbook, releasing its file handle.

    Args:
        workbook: The workbook to close.
    """
    workbook.close()


def resolve_range(workbook: Workbook, sheet: str, range: str) -> tuple[str, str]:
    """Resolve `range` into plain A1 notation on a specific sheet — either unchanged (already
    A1), or via a workbook-level defined name (PRD sec 7's named/defined-range support).

    Checking a workbook's *real* defined names can only happen once it's actually open (not a
    static-validation concern, Spec sec 5.4) — this is that runtime lookup, shared by every
    read action that accepts a `range`/`cell` field.

    Args:
        workbook: The workbook to resolve against.
        sheet: The worksheet `range` is scoped to when it's plain A1 notation. Ignored in
            favor of the defined name's own sheet when `range` is a defined name instead.
        range: An A1-style cell/range, or a workbook-level defined name.

    Returns:
        (resolved_sheet, a1_range).

    Raises:
        ValueError: If `range` is a defined name that resolves to more than one contiguous
            area (multi-area named ranges aren't supported) — raised here rather than left to
            fail confusingly further down in cell indexing.
    """
    defined_name = workbook.defined_names.get(range)
    if defined_name is None:
        return sheet, range
    destinations = list(defined_name.destinations)
    if len(destinations) != 1:
        raise ValueError(
            f'Named range "{range}" must resolve to exactly one area '
            f"(found {len(destinations)})."
        )
    dest_sheet, dest_range = destinations[0]
    return dest_sheet, dest_range.replace("$", "")


def read_range(workbook: Workbook, sheet: str, range: str) -> Any:
    """Read a cell or range of cells.

    Args:
        workbook: The workbook to read from.
        sheet: Worksheet name.
        range: An A1-style cell (e.g. "B2") or range (e.g. "A1:D50"), or a workbook-level
            defined name (resolved via `resolve_range`, which wins over `sheet` if the name
            points at a different sheet).

    Returns:
        The cell's value for a single cell, or a 2D list of row values for a range.

    Raises:
        ValueError: If `range` isn't valid A1 notation and isn't a real defined name either,
            or is a defined name spanning more than one area.
    """
    resolved_sheet, resolved_range = resolve_range(workbook, sheet, range)
    selection = workbook[resolved_sheet][resolved_range]
    if isinstance(selection, _SingleCell):
        return selection.value
    return [[cell.value for cell in row] for row in selection]


def resolve_sheet_names(
    workbook: Workbook, sheet: str | list[str] | dict[str, str]
) -> list[str]:
    """Resolve a `read_range`-style sheet spec into an explicit list of worksheet names.

    PRD sec 7's `read_range` row: an explicit list is the real multi-sheet mechanism;
    "all" and `{"matching": ...}` are both authoring sugar that expand to an explicit list
    before execution — one code path underneath either way.

    Args:
        workbook: The workbook whose sheets are being selected.
        sheet: A single sheet name, an explicit list of names, the literal string "all"
            (every sheet in the workbook), or `{"matching": <regex>}` (every sheet whose name
            matches the regex, via `re.search` — same convention as `find_row`/
            `find_headers_row`'s `patterns`).

    Returns:
        The resolved sheet names, in workbook order for "all"/`matching`.
    """
    if isinstance(sheet, list):
        return sheet
    if isinstance(sheet, dict):
        pattern = re.compile(sheet["matching"])
        return [name for name in workbook.sheetnames if pattern.search(name)]
    if sheet == "all":
        return list(workbook.sheetnames)
    return [sheet]


def write_cell(workbook: Workbook, sheet: str, cell: str, value: Any) -> None:
    """Write a value to a single cell.

    A value starting with "=" is stored as a formula automatically — openpyxl's normal
    behavior, not special-cased here (PRD sec 6.5's "formulas need no dedicated action").

    Args:
        workbook: The workbook to write to.
        sheet: Worksheet name.
        cell: An A1-style cell reference (e.g. "B2").
        value: The value to write.
    """
    workbook[sheet][cell] = value


def write_range(
    workbook: Workbook, sheet: str, range: str, values: list[list[Any]]
) -> None:
    """Write a 2D block of values, anchored at the top-left cell of `range`.

    The block's size comes from `values`, not from parsing the full extent of `range` — so
    `range` only needs its top-left cell to be correct (PRD sec 11 item 8).

    Args:
        workbook: The workbook to write to.
        sheet: Worksheet name.
        range: An A1-style cell or range — only the top-left cell is used as the anchor.
        values: A 2D list of row values to write.
    """
    anchor = workbook[sheet][range.split(":")[0]]
    start_row, start_col = anchor.row, anchor.column
    for row_offset, row_values in enumerate(values):
        for col_offset, value in enumerate(row_values):
            workbook[sheet].cell(
                row=start_row + row_offset, column=start_col + col_offset, value=value
            )


def set_column_width(
    workbook: Workbook, sheet: str, columns: str, width: float | Literal["autofit"]
) -> None:
    """Set the width of a column or range of columns.

    "autofit" approximates Excel's real autofit (which needs a rendering engine openpyxl
    doesn't have) by sizing to the longest value currently in each column, plus padding.

    Args:
        workbook: The workbook to modify.
        sheet: Worksheet name.
        columns: A single column letter (e.g. "B") or range (e.g. "A:C").
        width: An explicit width, or "autofit".
    """
    worksheet = workbook[sheet]
    start_letter, _, end_letter = columns.partition(":")
    end_letter = end_letter or start_letter
    for idx in range(
        column_index_from_string(start_letter), column_index_from_string(end_letter) + 1
    ):
        letter = get_column_letter(idx)
        if width == "autofit":
            lengths = [
                len(str(cell.value))
                for cell in worksheet[letter]
                if cell.value is not None
            ]
            worksheet.column_dimensions[letter].width = max(lengths, default=8) + 2
        else:
            worksheet.column_dimensions[letter].width = width


def create_sheet(workbook: Workbook, name: str, index: int | None = None) -> None:
    """Add a new, empty worksheet to the workbook.

    Args:
        workbook: The workbook to modify.
        name: Name for the new sheet.
        index: Position to insert at (0-based). Appended at the end if omitted.

    Raises:
        ValueError: If a sheet named `name` already exists.
    """
    if name in workbook.sheetnames:
        raise ValueError(f'A sheet named "{name}" already exists.')
    workbook.create_sheet(name, index)


def rename_sheet(workbook: Workbook, sheet: str, new_name: str) -> None:
    """Rename an existing worksheet.

    Args:
        workbook: The workbook to modify.
        sheet: Current worksheet name.
        new_name: New name for the worksheet.
    """
    workbook[sheet].title = new_name


def delete_sheet(workbook: Workbook, sheet: str) -> None:
    """Remove a worksheet from the workbook.

    Args:
        workbook: The workbook to modify.
        sheet: Name of the worksheet to remove.

    Raises:
        ValueError: If `sheet` is the workbook's only remaining sheet — a workbook can't have
            zero sheets, and openpyxl itself would leave the file in a broken state rather
            than raise, so this is checked explicitly here.
    """
    if len(workbook.sheetnames) == 1:
        raise ValueError(
            f'Cannot delete "{sheet}" — it is the only sheet left in the workbook.'
        )
    del workbook[sheet]


def insert_range(
    workbook: Workbook,
    sheet: str,
    at: str,
    direction: Literal["rows", "columns"] | None = None,
    header: dict[str, Any] | None = None,
) -> None:
    """Insert a whole row or whole column, shifting existing content.

    Args:
        workbook: The workbook to modify.
        sheet: Worksheet name.
        at: A whole-column reference (e.g. "C:C") or whole-row reference (e.g. "5:5").
        direction: Unused for whole-row/whole-column inserts — the direction is unambiguous
            from `at` itself (PRD sec 11 item 12). Accepted so partial-range calls (below)
            still reach the clear error instead of an unrelated TypeError.
        header: `{"row": int, "text": str}` — only meaningful for a column insert.

    Raises:
        NotImplementedError: If `at` is a partial range. Whole-row/whole-column insert is
            native to openpyxl and cheap; a true partial-range insert needs hand-rolled
            cell-shifting logic not built yet (PRD sec 11 item 12's flagged cost) — this must
            fail clearly rather than silently do the wrong thing.
    """
    worksheet = workbook[sheet]
    column_match = _WHOLE_COLUMN_RE.match(at)
    row_match = _WHOLE_ROW_RE.match(at)
    if column_match:
        idx = column_index_from_string(column_match.group(1))
        worksheet.insert_cols(idx)
        if header:
            worksheet.cell(row=header["row"], column=idx, value=header["text"])
    elif row_match:
        worksheet.insert_rows(int(row_match.group(1)))
    else:
        raise NotImplementedError(
            f'insert_range: partial range "{at}" is not supported yet — only whole-row '
            '("5:5") or whole-column ("C:C") ranges are built.'
        )


def copy_range(
    source_workbook: Workbook,
    source_sheet: str,
    source_range: str | None,
    target_workbook: Workbook,
    target_sheet: str,
    target_range: str,
) -> None:
    """Copy a range — or, if `source_range` is None, the whole used area of the sheet — from
    one workbook to another, anchored at `target_range`'s top-left cell.

    Args:
        source_workbook: The workbook to copy from.
        source_sheet: Source worksheet name.
        source_range: An A1-style range, or None to copy the whole sheet.
        target_workbook: The workbook to copy into.
        target_sheet: Target worksheet name.
        target_range: Where to start writing — only the top-left cell is used.
    """
    source_worksheet = source_workbook[source_sheet]
    if source_range is None:
        values = [[cell.value for cell in row] for row in source_worksheet.iter_rows()]
    else:
        selection = source_worksheet[source_range]
        values = (
            [[selection.value]]
            if isinstance(selection, _SingleCell)
            else [[cell.value for cell in row] for row in selection]
        )
    write_range(target_workbook, target_sheet, target_range, values)


def find_headers_row(
    workbook: Workbook, sheet: str, search_range: str, patterns: list[str]
) -> tuple[int, dict[str, str]] | None:
    """Find the row within `search_range` where every pattern matches some cell in that row.

    Args:
        workbook: The workbook to search.
        sheet: Worksheet name.
        search_range: An A1-style range to search within, or a workbook-level defined name.
        patterns: Regex patterns — every one must match a cell in a row for that row to count.

    Returns:
        `(row_number, {pattern: column_letter})` for the first matching row, or None.

    Raises:
        ValueError: If `search_range` isn't valid A1 notation and isn't a real defined name
            either, or is a defined name spanning more than one area.
    """
    resolved_sheet, resolved_range = resolve_range(workbook, sheet, search_range)
    selection = workbook[resolved_sheet][resolved_range]
    rows = [(selection,)] if isinstance(selection, _SingleCell) else selection
    for row in rows:
        matches: dict[str, str] = {}
        for pattern in patterns:
            for cell in row:
                if cell.value is not None and re.search(pattern, str(cell.value)):
                    matches[pattern] = get_column_letter(cell.column)
                    break
        if len(matches) == len(patterns):
            return row[0].row, matches
    return None


def find_row(
    workbook: Workbook,
    sheet: str,
    column: str,
    search_value: Any,
    header_row: int | None = None,
) -> int | None:
    """Find the row number where `column` equals `search_value`.

    Args:
        workbook: The workbook to search.
        sheet: Worksheet name.
        column: A column letter (e.g. "B").
        search_value: The value to match, by equality.
        header_row: If given, search starts on the row after it.

    Returns:
        The matching row number, or None if not found.
    """
    worksheet = workbook[sheet]
    col_idx = column_index_from_string(column)
    start_row = (header_row or 0) + 1
    for (cell,) in worksheet.iter_rows(
        min_row=start_row, min_col=col_idx, max_col=col_idx
    ):
        if cell.value == search_value:
            return cell.row
    return None


def find_column(
    workbook: Workbook, sheet: str, header_row: int, pattern: str
) -> str | None:
    """Find the column letter whose header (in `header_row`) matches `pattern`.

    Args:
        workbook: The workbook to search.
        sheet: Worksheet name.
        header_row: The row number containing headers.
        pattern: A regex pattern to match against each header cell's value.

    Returns:
        The matching column letter, or None if not found.
    """
    for cell in workbook[sheet][header_row]:
        if cell.value is not None and re.search(pattern, str(cell.value)):
            return get_column_letter(cell.column)
    return None


def find_columns(
    workbook: Workbook, sheet: str, header_row: int, patterns: dict[str, str]
) -> dict[str, str]:
    """Find multiple named columns by header pattern in one call.

    Args:
        workbook: The workbook to search.
        sheet: Worksheet name.
        header_row: The row number containing headers.
        patterns: Logical name to regex pattern.

    Returns:
        Logical name to column letter, for every pattern that matched. Names whose pattern
        didn't match anything are simply absent, not an error at this layer.
    """
    result: dict[str, str] = {}
    for name, pattern in patterns.items():
        column = find_column(workbook, sheet, header_row, pattern)
        if column is not None:
            result[name] = column
    return result


def read_properties(workbook: Workbook) -> dict[str, Any]:
    """Read a workbook's document properties (title, creator, etc.).

    Args:
        workbook: The workbook to read from.

    Returns:
        Mapping of property name to value, for every non-None standard property.
    """
    properties = workbook.properties
    return {
        name: value
        for name in vars(properties)
        if not name.startswith("_") and (value := getattr(properties, name)) is not None
    }


def read_cells(workbook: Workbook, sheet: str, cells: list[str]) -> dict[str, Any]:
    """Read a scattered list of specific cells.

    Args:
        workbook: The workbook to read from.
        sheet: Worksheet name.
        cells: A1-style cell references to read, or workbook-level defined names — each
            resolved independently (`resolve_range`), so a defined name pointing at a
            different sheet than `sheet` still works.

    Returns:
        Mapping of the original cell reference (as given in `cells`) to its value.

    Raises:
        ValueError: If a cell reference isn't valid A1 notation and isn't a real defined name
            either, or is a defined name spanning more than one area.
    """
    result: dict[str, Any] = {}
    for cell in cells:
        resolved_sheet, resolved_ref = resolve_range(workbook, sheet, cell)
        result[cell] = workbook[resolved_sheet][resolved_ref].value
    return result


def xlw_open_workbook(
    app: xw.App, path: str, mode: Literal["read_only", "read_write"]
) -> xw.Book:
    """Open an existing workbook in a live Excel App instance.

    Args:
        app: The App instance to open the workbook in — from `OwnedInstanceRegistry.spawn()`,
            never an unqualified `xw.Book(path)` that could attach to an instance not owned by
            this run (PRD sec 6.2.1).
        path: Path to an existing workbook file.
        mode: "read_only" opens without allowing writes; "read_write" allows them.

    Returns:
        The opened xlwings Book.

    Raises:
        FileNotFoundError: If path does not exist — xlwings itself raises this before any
            Apple Event/COM call is made, same contract as `open_workbook`.
    """
    # update_links=False, always: this project controls every external link refresh itself,
    # explicitly, via com_change_link/com_update_link (docs/recalc_and_link_refresh_plan.md).
    # Leaving Excel's own implicit link-update-on-open behavior enabled is not just redundant —
    # in a headless, invisible spawned App it can block indefinitely on a dialog nothing can
    # dismiss (confirmed empirically: reproduced exactly this hang, root-caused to this).
    logger.info("Opening workbook via Excel COM: %s (mode=%s)", path, mode)
    return app.books.open(path, read_only=(mode == "read_only"), update_links=False)


def xlw_close_workbook(book: xw.Book) -> None:
    """Close a workbook, without quitting the App instance it belongs to.

    Args:
        book: The workbook to close.
    """
    logger.debug('Closing workbook "%s"', book.name)
    book.close()


def xlw_save_workbook(book: xw.Book) -> None:
    """Save a workbook in place, to whatever path it was opened/created at.

    Args:
        book: The workbook to save.
    """
    logger.info('Saving workbook via Excel COM: "%s"', book.name)
    book.save()


# --- Copy (PRD sec 7's `copy` action) -------------------------------------------------------
#
# Genuinely COM-only, not just xlwings' portable API: Excel's own Copy/paste is what preserves
# formulas, formatting, and other cell properties across the copy — the file-backend
# `copy_range` above is deliberately value-only (openpyxl has no live calculation engine to
# re-anchor a copied formula's relative references against), which was fine for a first cut
# but not a faithful "copy" for anything containing formulas or formatting.


def com_copy_range(
    source_book: xw.Book,
    source_sheet: str,
    source_range: str | None,
    target_book: xw.Book,
    target_sheet: str,
    target_range: str,
) -> None:
    """Copy a range — or, if `source_range` is None, the whole used range of the sheet — from
    one live workbook to another, via Excel's own Copy, anchored at `target_range`'s top-left
    cell.

    Args:
        source_book: The live workbook to copy from.
        source_sheet: Source worksheet name.
        source_range: An A1-style range, or None to copy the whole used range of the sheet.
        target_book: The live workbook to copy into.
        target_sheet: Target worksheet name.
        target_range: Where to start pasting — only the top-left cell is used.
    """
    source_worksheet = source_book.sheets[source_sheet]
    source_selection = (
        source_worksheet.used_range
        if source_range is None
        else source_worksheet.range(source_range)
    )
    target_anchor = target_book.sheets[target_sheet].range(target_range.split(":")[0])
    logger.info(
        'Copying "%s" of "%s" to "%s" of "%s" via Excel COM',
        source_range or "<used range>",
        source_book.name,
        target_range,
        target_book.name,
    )
    source_selection.api.Copy(Destination=target_anchor.api)


# --- Recalculation (PRD sec 7's `recalculate` action) --------------------------------------
#
# `app.calculate()` is xlwings' own portable call — the only one of this group that isn't
# `com_`-prefixed, since it isn't reaching through `.api` at all. Everything else here needs
# per-workbook/per-sheet granularity or the app-wide "full"/"full_rebuild" force-refresh modes,
# neither of which xlwings exposes on its own portable surface — hence the raw COM object.

# COM's `Application.CalculationState` enum (xlCalculationState) — not wrapped by xlwings,
# these are the raw, stable Excel constants.
_XL_CALCULATION_DONE = 0  # xlDone
_DEFAULT_CALCULATION_WAIT_TIMEOUT_SECONDS = 300.0
_CALCULATION_POLL_INTERVAL_SECONDS = 0.25


def xlw_calculate_all(app: xw.App) -> None:
    """Recalculate every open workbook in `app` (xlwings' own portable API).

    Args:
        app: The App instance whose open workbooks should recalculate.
    """
    logger.info("Recalculating all open workbooks")
    app.calculate()


def com_calculate_workbook(book: xw.Book) -> None:
    """Recalculate every sheet in a single workbook, via COM.

    There is no `Workbook.Calculate` in the Excel object model at all — only
    `Application.Calculate` (every open workbook) and `Worksheet.Calculate` (one sheet).
    Workbook-scoped recalculation is built here as looping `Worksheet.Calculate` over every
    sheet in `book` — confirmed necessary the hard way: `book.api.Calculate()` itself raises
    `AttributeError`/COM "Unknown name" since the member genuinely doesn't exist.

    Args:
        book: The workbook to recalculate.
    """
    logger.info('Recalculating workbook "%s"', book.name)
    for sheet in book.sheets:
        sheet.api.Calculate()


def com_calculate_sheet(book: xw.Book, sheet: str) -> None:
    """Recalculate a single worksheet, via COM (no portable xlwings equivalent exists at this
    granularity).

    Args:
        book: The workbook containing the sheet.
        sheet: Worksheet name.
    """
    logger.info('Recalculating sheet "%s" in workbook "%s"', sheet, book.name)
    book.sheets[sheet].api.Calculate()


def com_calculate_full(app: xw.App) -> None:
    """Force a full recalculation of every open workbook in `app`, including cells Excel
    wouldn't otherwise consider dirty — always application-wide, there is no per-workbook
    equivalent in COM.

    Args:
        app: The App instance to fully recalculate.
    """
    logger.info("Forcing full recalculation of all open workbooks")
    app.api.CalculateFull()


def com_calculate_full_rebuild(app: xw.App) -> None:
    """Force a full rebuild recalculation (rechecks dependency trees too, not just marks-dirty
    cells) of every open workbook in `app` — always application-wide, there is no per-workbook
    equivalent in COM.

    Args:
        app: The App instance to fully rebuild-recalculate.
    """
    logger.info("Forcing full rebuild recalculation of all open workbooks")
    app.api.CalculateFullRebuild()


def com_wait_until_calculation_done(
    app: xw.App, timeout: float = _DEFAULT_CALCULATION_WAIT_TIMEOUT_SECONDS
) -> None:
    """Block until `app`'s calculation has actually finished.

    Excel can return control from a Calculate()/CalculateFull() call before a large model has
    actually finished recalculating — saving before that completes would silently persist
    stale values. Polls `Application.CalculationState` rather than trusting the call to be
    synchronous.

    Args:
        app: The App instance to wait on.
        timeout: Maximum seconds to wait before giving up.

    Raises:
        TimeoutError: If calculation hasn't finished within `timeout`.
    """
    deadline = time.monotonic() + timeout
    logger.info("Waiting for Excel calculation to finish (timeout=%ss)", timeout)
    while app.api.CalculationState != _XL_CALCULATION_DONE:
        if time.monotonic() >= deadline:
            raise TimeoutError(
                f"Excel calculation did not finish within {timeout} seconds"
            )
        time.sleep(_CALCULATION_POLL_INTERVAL_SECONDS)
    logger.info("Excel calculation finished")


# --- Link management (external link repointing, docs/recalc_and_link_refresh_plan.md) ------

_XL_LINK_TYPE_EXCEL_LINKS = (
    1  # xlLinkTypeExcelLinks — the only link type this project handles
)


def com_link_sources(book: xw.Book) -> list[str]:
    """List every external Excel-workbook link source a workbook currently has.

    Args:
        book: The workbook to inspect.

    Returns:
        Each link's current source, in whatever form Excel currently has it stored (a bare
        filename for an unsaved/same-folder link, an absolute/UNC path otherwise). Empty if
        the workbook has no external Excel-workbook links.
    """
    sources = book.api.LinkSources(_XL_LINK_TYPE_EXCEL_LINKS)
    result = list(sources) if sources else []
    logger.debug('Workbook "%s" external link sources: %s', book.name, result)
    return result


def com_change_link(book: xw.Book, name: str, new_name: str) -> None:
    """Repoint one of a workbook's external links to a new target.

    `name` must match one of `com_link_sources(book)`'s current strings exactly — Excel
    matches a link to an open workbook by that stored string, not by filename alone. If
    `new_name` exists on disk, Excel immediately re-evaluates dependent cells against its
    current content (confirmed empirically); if it does not exist, dependent cells are
    blanked immediately instead. See docs/recalc_and_link_refresh_plan.md.

    Args:
        book: The workbook whose link is being repointed.
        name: The link's current source, exactly as returned by `com_link_sources`.
        new_name: The new target path.
    """
    logger.info(
        'Repointing link "%s" -> "%s" in workbook "%s"', name, new_name, book.name
    )
    book.api.ChangeLink(Name=name, NewName=new_name, Type=_XL_LINK_TYPE_EXCEL_LINKS)


def com_update_link(book: xw.Book, name: str) -> None:
    """Force a fresh read of one of a workbook's external links from disk.

    Works even if the source workbook was never opened by this process at all — confirmed
    empirically to read directly from the file on disk (docs/recalc_and_link_refresh_plan.md).

    Args:
        book: The workbook whose link should be refreshed.
        name: The link's current source, exactly as returned by `com_link_sources`.
    """
    logger.debug('Updating link "%s" in workbook "%s" from disk', name, book.name)
    book.api.UpdateLink(Name=name, Type=_XL_LINK_TYPE_EXCEL_LINKS)


# --- xlwings — owned-instance tracking (PRD sec 6.2.1, Spec sec 3.1) ----------------------


class OwnedInstanceRegistry:
    """Tracks Excel App instances this run has spawned itself, so it only ever closes those.

    xlwings (and the underlying Excel COM/Apple Event API) will happily attach to whatever
    Excel instance is already running rather than spawning its own — dangerous when more than
    one automation run, or a real user's own Excel session, can be active at once. Every
    instance this registry hands out comes from a fresh `xw.App(...)` call, never an
    unqualified lookup like `xw.apps.active` or a bare `xw.Book("name.xlsx")` that could bind
    to an instance it doesn't own.
    """

    def __init__(self) -> None:
        self._owned: dict[int, xw.App] = {}

    @property
    def pids(self) -> tuple[int, ...]:
        """Process IDs of every instance currently owned — the run's audit trail of what it
        spawned (PRD sec 6.2.1's "records that instance's process ID against the run").
        """
        return tuple(self._owned)

    def spawn(self, visible: bool = False) -> xw.App:
        """Spawn a brand-new, dedicated Excel App instance and start tracking it.

        Args:
            visible: Whether the spawned Excel window is shown. Defaults to hidden, the normal
                automation case; a visible instance is for a deliberately different use (e.g.
                the parked "replay nice" idea, PRD sec 12), not the default run path.

        Returns:
            The newly spawned App.
        """
        # add_book=True, not False: found via a real, reproducible Windows-only bug (~50% of
        # runs) where a bookless instance's own `App.quit()` handle would hang indefinitely.
        # A bookless xw.App() is never registered in `xw.apps` at all (confirmed directly:
        # spawning two bookless instances left `xw.apps` empty), and closing an instance
        # xlwings itself doesn't know about is what caused the intermittent hang. With
        # add_book=True the instance gets an initial "Book1" and registers immediately —
        # verified reliable across repeated spawn/quit cycles. Matches the always-add_book
        # default of the older, previously-reliable reference implementation this project
        # replaces (Risk Demo's excel_core.py).
        app = xw.App(visible=visible, add_book=True)
        # Suppress every alert/prompt dialog by default: an automation run has no user to
        # answer them, and in a headless/invisible instance an unanswered modal blocks
        # forever (confirmed empirically — reproduced exactly this hang, root-caused to
        # missing alert suppression). This project never wants interactive prompts, so these
        # are unconditional defaults, not per-call opt-ins.
        app.display_alerts = False
        app.api.AskToUpdateLinks = False
        self._owned[app.pid] = app
        logger.info("Spawned owned Excel instance (pid=%d)", app.pid)
        return app

    def close_owned(self) -> None:
        """Quit every owned instance, attempting all of them even if some fail.

        Raises:
            ExceptionGroup: If one or more instances failed to quit. Every instance still gets
                a quit attempt regardless (PRD sec 6.3's crash-safety requirement), mirroring
                `engine.SessionManager.close_all()` — this isn't a defensive catch-and-ignore,
                every failure is still surfaced, just after giving every other instance a
                chance to close too.
        """
        errors: list[Exception] = []
        for pid, app in tuple(self._owned.items()):
            logger.info("Closing owned Excel instance (pid=%d)", pid)
            try:
                # Quit through a fresh xw.apps[pid] lookup, not the stored App object directly
                # — this is the same fix as `add_book=True` above: the instance is registered
                # in xw.apps now, so this is the reliable, verified-repeatedly path. Fall back
                # to the stored object's own quit() if it's not in xw.apps (e.g. a test double,
                # or an instance closed by some other means already) — that fallback path can
                # still raise, and is still caught and surfaced below like any other failure.
                try:
                    xw.apps[pid].quit()
                except KeyError:
                    app.quit()
            except Exception as exc:  # noqa: BLE001 - intentional, see docstring
                errors.append(exc)
        self._owned.clear()
        if errors:
            raise ExceptionGroup(
                "failed to quit one or more owned Excel instances", errors
            )
