"""Run-preparation and run-state layer: action discovery, session management, the
scratch-copy execution model, and both validation tiers. See Spec sec 5.
"""

import difflib
import inspect
import logging
import re
import shutil
import types as pytypes
import typing
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from types import ModuleType
from typing import Any, Literal
from urllib.parse import unquote, urlparse
from xml.etree import ElementTree

import openpyxl

from excel_runner import backends
from excel_runner.core import (
    ACTION_CAPABILITIES,
    ACTION_WRITES,
    ActionExecutionError,
    ActionResult,
    ErrorDetail,
    Step,
    ValidationError,
    WorkbookRef,
    WorkbookSession,
    Workflow,
    is_whole_template_expression,
)

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ActionSpec:
    """A discovered action: its name, callable, capability, description, and parameter schema.

    Args:
        name: The action's name, matching an `action:` field in a workflow step.
        fn: The action function itself.
        capability: Which backend this action needs. "depends_on_param" is a named, single
            exception (PRD sec 7's `read_metadata`) — not a general mechanism.
        description: The action's docstring, first line only — what a future agent-tool
            wrapper would show as the tool's description (PRD sec 6.1's "close to free" tool
            schema generation). Added after the fact: the field didn't exist until §6.3's
            `list_actions()` was being built and needed it.
        param_schema: `{"properties": {name: {"type": ...}}, "required": [...]}`, derived from
            `fn`'s signature (excluding `session`).
        writes: Whether this action mutates the workbook it's given — declared via
            `writes=True` on the same `@file_action`/`@xlw_action`/`@com_action` decorator
            that registers `capability` (Spec sec 5.4), not a separate hardcoded list. Used by
            `plan()` below to infer read_only vs. read_write per workbook.
    """

    name: str
    fn: Callable[..., ActionResult]
    capability: Literal["file", "xlw", "com", "depends_on_param", "none"]
    description: str
    param_schema: dict[str, Any]
    writes: bool = False


def _generate_param_schema(fn: Callable[..., ActionResult]) -> dict[str, Any]:
    """Build a param schema from an action function's signature, excluding `session`/
    `step_outputs` — both are injected by the runner, never a real YAML field."""
    signature = inspect.signature(fn)
    properties: dict[str, Any] = {}
    required: list[str] = []
    for param_name, param in signature.parameters.items():
        if param_name in ("session", "step_outputs"):
            continue
        properties[param_name] = {"type": param.annotation}
        if param.default is inspect.Parameter.empty:
            required.append(param_name)
    return {"properties": properties, "required": required}


def discover_actions(module: ModuleType) -> dict[str, ActionSpec]:
    """Scan a module for capability-tagged action functions and build the registry.

    Args:
        module: The module to scan — normally `excel_runner.actions`.

    Returns:
        Mapping of action name to `ActionSpec`.
    """
    registry: dict[str, ActionSpec] = {}
    for name, fn in inspect.getmembers(module, inspect.isfunction):
        capability = ACTION_CAPABILITIES.get(name)
        if capability is None:
            continue
        registry[name] = ActionSpec(
            name=name,
            fn=fn,
            capability=capability,
            description=_first_line(inspect.getdoc(fn) or ""),
            param_schema=_generate_param_schema(fn),
            writes=ACTION_WRITES.get(name, False),
        )
    return registry


def _first_line(docstring: str) -> str:
    return docstring.split("\n", 1)[0].strip()


# --- R4 link commit ordering (docs/recalc_and_link_refresh_plan.md R5-R7) ------------------


_UNC_OR_DRIVE_ABSOLUTE = re.compile(r"^([a-zA-Z]:[\\/]|\\\\|//)")
_DRIVELESS_ROOTED = re.compile(r"^/(?!/)")


def classify_link_target(
    target: str,
) -> Literal["same_folder", "relative_subpath", "absolute"]:
    """Classify a raw external-link Target string (as stored in an xlsx's
    `externalLinks/_rels/*.rels`) per plan doc sec 2's R1/R2/R3-R4 categories.

    Args:
        target: The raw `Target` attribute value from an external-link relationship.

    Returns:
        `"same_folder"` (R1 — a bare filename, no path separator at all), `"relative_subpath"`
        (R2, backlog/unsupported — relative but not same-folder, e.g. `"other/x.xlsx"` or
        `"../x.xlsx"`), or `"absolute"` (R3/R4 — a drive-letter path, a UNC path, a `file://`
        URI, or a drive-omitted rooted path like `"/Users/x/target.xlsx"` — what real Excel
        actually writes for a link to a file on the *same drive* as the linking workbook,
        confirmed via a real Excel-authored fixture; `resolve_link_target`'s existing
        parent-join fallback already resolves this form correctly, re-rooted onto the linking
        workbook's own drive).
    """
    if target.startswith("file://"):
        return "absolute"
    if _UNC_OR_DRIVE_ABSOLUTE.match(target):
        return "absolute"
    if _DRIVELESS_ROOTED.match(target):
        return "absolute"
    if "/" in target or "\\" in target:
        return "relative_subpath"
    return "same_folder"


def resolve_link_target(target: str, linking_workbook_path: Path) -> Path:
    """Resolve a raw external-link Target string to a real, absolute filesystem path.

    Args:
        target: The raw `Target` attribute value from an external-link relationship — same
            or relative form resolved against `linking_workbook_path`'s folder, or an
            already-absolute drive/UNC path, or a `file://` URI, returned as-is (converted).
        linking_workbook_path: Real path of the workbook the link was found in — same/relative
            targets are resolved relative to its parent folder.

    Returns:
        The resolved, absolute `Path`.
    """
    if target.startswith("file://"):
        return Path(unquote(urlparse(target).path).lstrip("/")).resolve()
    if _UNC_OR_DRIVE_ABSOLUTE.match(target):
        return Path(target).resolve()
    return (linking_workbook_path.parent / target).resolve()


def scan_external_link_targets(path: Path) -> list[str]:
    """Read every external-link Target string out of a real xlsx file, with no Excel/COM
    involved at all — a plain zipfile/XML read, safe to call during planning.

    Args:
        path: Real path of an existing `.xlsx` file.

    Returns:
        Raw `Target` strings, one per external link found (order as stored in the file).
        Empty list if the workbook has no external links.
    """
    targets: list[str] = []
    with zipfile.ZipFile(path) as archive:
        rels_names = sorted(
            name
            for name in archive.namelist()
            if name.startswith("xl/externalLinks/_rels/") and name.endswith(".rels")
        )
        for rels_name in rels_names:
            root = ElementTree.fromstring(archive.read(rels_name))
            for relationship in root:
                if relationship.get("TargetMode") == "External":
                    target = relationship.get("Target")
                    if target is not None:
                        targets.append(target)
    return targets


def discover_write_intent_link_graph(
    workbook_paths: dict[str, Path], write_intent: set[str]
) -> dict[str, set[str]]:
    """Build the R4 `link_targets` graph `compute_link_commit_order()` consumes, by scanning
    every write-intent workbook's real, on-disk file for `"absolute"`-classified external
    links (R3/R4) that resolve to another *write-intent* declared workbook.

    R1 (same-folder) and R2 (relative-subpath) links are not R4 concerns — ignored here. An
    absolute link that resolves to a declared workbook which is *not* write-intent is R3 (leave
    untouched) — also ignored, since nothing about that target changes this run. An absolute
    link that doesn't resolve to any declared workbook at all (something outside this run
    entirely) is likewise ignored.

    Args:
        workbook_paths: Every declared workbook's logical name to its real, on-disk path.
        write_intent: Logical names of workbooks that will be modified this run (`plan()`'s
            `"read_write"` workbooks).

    Returns:
        Logical name -> set of other write-intent workbook names it R4-links to. Every name in
        `write_intent` is present as a key (empty set if it has no R4 links), matching
        `compute_link_commit_order()`'s expected input shape.
    """
    resolved_paths = {
        name: path.resolve() for name, path in workbook_paths.items() if path.exists()
    }
    graph: dict[str, set[str]] = {name: set() for name in write_intent}
    for name in write_intent:
        path = workbook_paths.get(name)
        if path is None or not path.exists():
            continue
        for target in scan_external_link_targets(path):
            if classify_link_target(target) != "absolute":
                continue
            resolved_target = resolve_link_target(target, path)
            for other_name, other_path in resolved_paths.items():
                if other_name != name and other_path == resolved_target:
                    if other_name in write_intent:
                        graph[name].add(other_name)
                    break
    return graph


def compute_link_commit_order(link_targets: dict[str, set[str]]) -> list[str]:
    """Topologically order write-intent workbooks so each is committed after every workbook
    its R4 links point to (R5).

    Args:
        link_targets: For every write-intent workbook, the set of other write-intent
            workbooks it has an R4 (absolute/UNC, to-be-modified) link to. Workbooks with no
            outbound R4 links must still be present as keys, mapped to an empty set.

    Returns:
        Workbook names in commit order.

    Raises:
        ValidationError: On a cyclical R4 link between two workbooks (R6), or a link chain
            deeper than one hop (R7) — a workbook that is both an R4 link's target and the
            source of an R4 link to a different workbook.
    """
    incoming: dict[str, set[str]] = {name: set() for name in link_targets}
    for name, targets in link_targets.items():
        for target in targets:
            incoming.setdefault(target, set()).add(name)
    for name, targets in link_targets.items():
        sources = incoming.get(name, set())
        # A clean two-way pair (A->B, B->A) is a cycle (R6), caught below by the topological
        # sort itself never finding a ready node. Only reject here when this workbook's
        # outbound target(s) go beyond exactly mirroring back its own incoming source(s) —
        # that's a genuine chain (R7), not just a two-node cycle.
        if targets and sources and targets != sources:
            raise ValidationError(
                ErrorDetail(
                    message=(
                        f'Workbook "{name}" is both the target of an R4 link and has its own '
                        "outbound R4 link to a different workbook — link chains beyond one "
                        "hop are not supported."
                    ),
                    technical_reason=f"link_targets={link_targets!r}",
                )
            )

    order: list[str] = []
    remaining = {name: set(targets) for name, targets in link_targets.items()}
    while remaining:
        ready = sorted(name for name, targets in remaining.items() if not targets)
        if not ready:
            raise ValidationError(
                ErrorDetail(
                    message=(
                        "Cyclical R4 links detected between: "
                        f"{', '.join(sorted(remaining))} — no valid commit order exists."
                    ),
                    technical_reason=f"remaining={remaining!r}",
                )
            )
        order.extend(ready)
        for name in ready:
            del remaining[name]
        for targets in remaining.values():
            targets.difference_update(ready)
    return order


# --- Scratch-copy execution model (Spec sec 5.3, PRD sec 6.3.1;
# docs/recalc_and_link_refresh_plan.md sec 1/3) --------------------------------------------


class ScratchManager:
    """Stages workbooks that will be written to into a scratch directory, and commits them
    back to their real path, only on success. Operates on plain file paths — no knowledge of
    openpyxl/xlwings, so file-backend and (later) COM-backend sessions stage and commit through
    the same code path (PRD sec 6.3.1).

    Scratch has two subfolders (plan doc sec 1), not one flat directory:
    `working_dir/scratch/working/` holds every staged workbook's live copy, named with its
    *original real basename* (not its logical/YAML name) — needed so a same-folder (R1)
    external link between two staged workbooks still resolves once both sit in this same
    folder. `working_dir/scratch/originals/` holds a pre-edit backup, made only for
    write-intent workbooks, for commit-time rollback safety and as an untouched reference
    (sec 1.4) — a purely defensive copy, never opened or read by this class itself.

    Args:
        working_dir: The run's working directory (PRD sec 6.3.4) — scratch subfolders are
            created lazily on first `stage()` call, never just by constructing a
            `ScratchManager`.
    """

    def __init__(self, working_dir: Path) -> None:
        self._working_subdir = working_dir / "scratch" / "working"
        self._originals_dir = working_dir / "scratch" / "originals"
        self._staged: dict[str, tuple[Path, Path, bool]] = (
            {}
        )  # name -> (real_path, working_path, writes)
        self._backups: dict[str, Path] = (
            {}
        )  # name -> .bak path, only set during a commit_all()

    def stage(self, name: str, real_path: Path, writes: bool = True) -> Path:
        """Copy a workbook into `scratch/working/`, or reserve a path there for a new one.

        Args:
            name: The workbook's logical name.
            real_path: Its real file path. If it doesn't exist yet (a `create_if_missing`
                workbook), no copy happens — the caller creates the workbook directly at the
                returned working path instead.
            writes: Whether this workbook may be written to and needs committing back later.
                False for a read-only session (PRD sec 6.2.3's correction — staged too now, to
                avoid holding a handle open on the real file, but never committed since
                nothing about it ever changes). Only a `writes=True` workbook gets a
                `scratch/originals/` backup — a read-only copy never changes, so there's
                nothing to back up.

        Returns:
            The `scratch/working/` path to open/create the workbook at instead of `real_path`.

        Note:
            Two declared workbooks with the same real basename (in different real folders)
            would collide in `scratch/working/` — not handled specially; not expected in
            practice and not worth the extra complexity unless it actually comes up.
        """
        logger.info('Staging workbook "%s" into scratch: %s', name, real_path)
        self._working_subdir.mkdir(parents=True, exist_ok=True)
        working_path = self._working_subdir / real_path.name
        if real_path.exists():
            shutil.copy2(real_path, working_path)
            if writes:
                self._originals_dir.mkdir(parents=True, exist_ok=True)
                shutil.copy2(real_path, self._originals_dir / real_path.name)
        else:
            logger.debug(
                'Workbook "%s" has no real file yet — created on first write', name
            )
        self._staged[name] = (real_path, working_path, writes)
        return working_path

    def working_path(self, name: str) -> Path:
        """The `scratch/working/` path a staged workbook was given, as returned by `stage()`.

        Args:
            name: The workbook's logical name, as passed to `stage()`.
        """
        return self._staged[name][1]

    def real_path(self, name: str) -> Path:
        """The real, on-disk path a staged workbook will eventually be committed back to.

        Args:
            name: The workbook's logical name, as passed to `stage()`.
        """
        return self._staged[name][0]

    def commit(self, name: str) -> None:
        """Commit one staged workbook's working content back to its real path.

        Copy-based, not rename-based (plan doc sec 3.2): if `real_path` already exists, it is
        first *copied* (never moved/deleted) to a `.bak` sibling, so the original stays
        recoverable even if the following overwrite is interrupted. The working copy is then
        copied onto `real_path`. The `.bak` is left in place here — `commit_all()` deletes
        every one only after every workbook in the batch has committed successfully, or copies
        it back to roll back on a later failure.

        Args:
            name: The workbook's logical name, as passed to `stage()`.
        """
        real_path, working_path, _ = self._staged[name]
        logger.info('Committing workbook "%s" to %s', name, real_path)
        real_path.parent.mkdir(parents=True, exist_ok=True)
        if real_path.exists():
            bak_path = real_path.with_name(real_path.name + ".bak")
            shutil.copy2(real_path, bak_path)
            self._backups[name] = bak_path
        shutil.copy2(working_path, real_path)

    def commit_all(
        self,
        order: list[str] | None = None,
        before_commit: Callable[[str], None] | None = None,
    ) -> None:
        """Commit every staged, write-intent workbook, rolling back on a later failure (PRD
        sec 6.3.3; plan doc sec 3). Read-only staged workbooks (`stage(..., writes=False)`) are
        skipped entirely — nothing about them ever changes, so there's nothing to commit back.

        No separate upfront precheck pass — each workbook's commit is attempted directly. If
        one fails, every workbook already committed *in this call* is rolled back (its `.bak`
        copied back over `real_path`, reverse order), and whether each individual rollback
        itself succeeded is recorded. A workbook whose rollback also fails needs a human — its
        `.bak` is deliberately left in place rather than deleted, so the original content is
        still recoverable from disk. On full success, every `.bak` created this call is deleted.

        Args:
            order: Commit order, as names (plan doc sec 3.1's R5 dependency order — a workbook
                that another one's R4 link points to must be committed first). Defaults to the
                natural write-intent staging order when there's no R4 link graph to respect.
            before_commit: Called with a workbook's name right before its own `commit()` —
                this is where `SessionManager` hooks in R4's commit-time link-revert-and-save
                (plan doc sec 3.2.1), so it runs while the workbook is still open and its link
                still points at its target's scratch copy, before that copy's content gets
                copied onto the workbook's own real path. Any exception it raises is treated
                exactly like a `commit()` failure — same rollback of already-committed
                workbooks.

        Raises:
            ActionExecutionError: If any workbook's commit (or `before_commit` hook) fails.
                The message names the workbook that failed, and — if any earlier workbook's
                rollback in this same call also failed — names which one(s) need manual
                attention.
        """
        names = (
            order
            if order is not None
            else [name for name, (_, _, writes) in self._staged.items() if writes]
        )
        logger.info("Committing %d workbook(s): %s", len(names), ", ".join(names))
        committed: list[str] = []
        for name in names:
            try:
                if before_commit is not None:
                    before_commit(name)
                self.commit(name)
                committed.append(name)
            except (
                Exception
            ) as exc:  # noqa: BLE001 - before_commit may raise a COM error too
                rollback_results = self._rollback(committed)
                needs_human = [n for n, ok in rollback_results.items() if not ok]
                rolled_back = [n for n, ok in rollback_results.items() if ok]
                message = f'Workbook "{name}" could not be committed: {exc}.'
                if rolled_back:
                    message += f" Rolled back: {', '.join(rolled_back)}."
                if needs_human:
                    message += (
                        f" MANUAL INTERVENTION NEEDED for: {', '.join(needs_human)} — "
                        "their .bak file(s) hold the original content."
                    )
                raise ActionExecutionError(
                    ErrorDetail(
                        message=message,
                        technical_reason=(
                            f"{type(exc).__name__}: {exc}; rollback_results={rollback_results!r}"
                        ),
                    )
                ) from exc
        for bak_path in self._backups.values():
            bak_path.unlink(missing_ok=True)
        self._backups.clear()
        logger.info("Commit complete: %d workbook(s) committed", len(names))

    def _rollback(self, committed_names: list[str]) -> dict[str, bool]:
        """Undo every already-committed workbook in `committed_names`, reverse order.

        Args:
            committed_names: Logical names committed so far in this `commit_all()` call.

        Returns:
            Logical name -> whether that workbook's rollback succeeded.
        """
        results: dict[str, bool] = {}
        for name in reversed(committed_names):
            real_path, _, _ = self._staged[name]
            bak_path = self._backups.get(name)
            try:
                real_path.unlink(missing_ok=True)
                if bak_path is not None:
                    shutil.copy2(bak_path, real_path)
                results[name] = True
            except OSError:
                results[name] = False
        return results


# --- Session management (Spec sec 5.2) ----------------------------------------------------


def _needed_backend(
    capability: Literal["file", "xlw", "com", "depends_on_param", "none"],
) -> Literal["file", "xlw"]:
    """Which `WorkbookSession.backend` a given action capability needs (PRD sec 6.2.2).

    `"com"` needs an `xlw`-backed session too — the action itself reaches deeper via xlwings'
    `.api`, so `SessionManager` never needs a distinct backend state for it. `"depends_on_param"`
    (`read_metadata`'s own runtime resolution, Spec sec 5.1) and `"none"` (control actions, PRD
    sec 6.9 — never call `get_or_open` at all, no `workbook:` field) can't be mapped to a
    concrete backend here.

    Raises:
        ActionExecutionError: For `"depends_on_param"` or `"none"` — not yet resolvable, or
            never actually reachable from here.
    """
    if capability == "file":
        return "file"
    if capability in ("xlw", "com"):
        return "xlw"
    raise ActionExecutionError(
        ErrorDetail(
            message=f'Action capability "{capability}" can\'t be resolved to a session backend yet.',
            technical_reason=(
                f"_needed_backend: unsupported capability {capability!r} (PRD sec 6.2.2/Spec sec 5.1)"
            ),
        )
    )


class SessionManager:
    """Tracks one WorkbookSession per logical workbook name for the duration of a run.

    Read/write mode is caller-specified, not statically inferred — tier-2 validation
    (Spec sec 5.4, not built yet) will compute that and hand it in later; for now the caller
    decides. Bidirectional backend switching (file <-> xlw mid-run, PRD sec 6.2.2) is built:
    `get_or_open` switches a session's backend in place (closing/reopening its handle) rather
    than raising, when a capability doesn't match a session's current backend.

    Every xlw/com-capability session opened by this manager, for every workbook, shares one
    lazily-spawned Excel instance (`self._owned_instances`, spawned on first need) — not one
    instance per workbook. Verified empirically: two workbooks opened via `app.books.open()`
    on the same `xw.App` land in the same Excel process (same `app.pid` on both), which is
    also what lets Excel resolve/recalculate live links between them. A per-workbook instance
    would defeat that.

    Args:
        workbooks: The workflow's `workbooks:` registry, name to WorkbookRef.
        scratch: Where read-write sessions get staged (PRD sec 6.3.1).
        link_targets: R4 link graph (`discover_write_intent_link_graph`'s output) — source
            workbook name to the set of other write-intent workbook names it has an R4
            (absolute/UNC) external link to. `None`/empty if there's no such link in this run,
            the common case — nothing extra happens then.
        commit_order: R4 commit order (`compute_link_commit_order`'s output over
            `link_targets`, computed once upfront so a cyclical/chained link (R6/R7) raises
            before any workbook is touched) — every write-intent workbook committed in this
            order instead of arbitrary staged order, so a workbook is always committed after
            every workbook its own R4 link(s) point to.
    """

    def __init__(
        self,
        workbooks: dict[str, WorkbookRef],
        scratch: ScratchManager,
        link_targets: dict[str, set[str]] | None = None,
        commit_order: list[str] | None = None,
    ) -> None:
        self._workbooks = workbooks
        self._scratch = scratch
        self._sessions: dict[str, WorkbookSession] = {}
        self._owned_instances = backends.OwnedInstanceRegistry()
        self._app: Any = None
        self._link_targets: dict[str, set[str]] = link_targets or {}
        self._link_sources: dict[str, set[str]] = {}
        for source, targets in self._link_targets.items():
            for target in targets:
                self._link_sources.setdefault(target, set()).add(source)
        self._commit_order = commit_order
        self._wired_r4_links: set[tuple[str, str]] = set()

    def _shared_app(self) -> Any:
        """Return the run's one shared, lazily-spawned Excel instance, spawning it on first
        need (PRD sec 6.2.1) — never one instance per workbook, see class docstring."""
        if self._app is None:
            self._app = self._owned_instances.spawn()
        return self._app

    def get_or_open(
        self,
        name: str,
        mode: Literal["read_only", "read_write"] = "read_write",
        capability: Literal["file", "xlw", "com", "depends_on_param", "none"] = "file",
    ) -> WorkbookSession:
        """Return the session for `name`, opening it on first reference.

        A `mode="read_write"` session is staged through the scratch-copy manager first (PRD
        sec 6.3.1) — real work happens on the scratch copy, never the original, until
        `commit_all()`. A `mode="read_only"` session opens directly against the real path.

        Args:
            name: The workbook's logical name, matching a key in the `workbooks:` registry.
            mode: Ignored if a session for `name` is already open — the mode it was first
                opened with sticks for the rest of the run.
            capability: The dispatching action's capability (Spec sec 5.1) — determines which
                backend (`_needed_backend`, PRD sec 6.2.2) this session must be on. Every
                session opens on the file backend today (bidirectional switching isn't built
                yet), so this only matters for detecting a mismatch, not yet for resolving one.

        Returns:
            The (possibly newly-opened, possibly just-switched) WorkbookSession, on the
            backend `capability` needs.

        Raises:
            ActionExecutionError: If `name` isn't in the registry, or its file doesn't exist
                and `create_if_missing` isn't set.
        """
        needed = _needed_backend(capability)
        if name in self._sessions:
            session = self._sessions[name]
            if session.backend != needed:
                self._switch_backend(session, needed)
            return session

        if name not in self._workbooks:
            raise ActionExecutionError(
                ErrorDetail(
                    message=f'Workbook "{name}" is not declared in the workbooks: registry.',
                    technical_reason=f"SessionManager.get_or_open: unknown workbook name {name!r}",
                )
            )
        ref = self._workbooks[name]
        session = (
            self._open_read_write(name, ref, needed)
            if mode == "read_write"
            else self._open_read_only(name, ref, needed)
        )
        self._sessions[name] = session
        self._wire_r4_links_touching(name)
        return session

    def _open_handle(
        self,
        scratch_path: Path,
        mode: Literal["read_only", "read_write"],
        backend: Literal["file", "xlw"],
    ) -> Any:
        if backend == "file":
            return backends.open_workbook(str(scratch_path), mode=mode)
        return backends.xlw_open_workbook(self._shared_app(), str(scratch_path), mode)

    def _open_read_write(
        self, name: str, ref: WorkbookRef, backend: Literal["file", "xlw"] = "file"
    ) -> WorkbookSession:
        real_path = Path(ref.file)
        scratch_path = self._scratch.stage(name, real_path)
        if not scratch_path.exists():
            self._create(ref, scratch_path)
        handle = self._open_handle(scratch_path, "read_write", backend)
        return WorkbookSession(
            name=name,
            backend=backend,
            handle=handle,
            path=str(scratch_path),
            mode="read_write",
            scratch_path=scratch_path,
        )

    def _open_read_only(
        self, name: str, ref: WorkbookRef, backend: Literal["file", "xlw"] = "file"
    ) -> WorkbookSession:
        real_path = Path(ref.file)
        scratch_path = self._scratch.stage(name, real_path, writes=False)
        if not scratch_path.exists():
            self._create(ref, scratch_path)
        handle = self._open_handle(scratch_path, "read_only", backend)
        return WorkbookSession(
            name=name,
            backend=backend,
            handle=handle,
            path=str(scratch_path),
            mode="read_only",
            scratch_path=scratch_path,
        )

    def _switch_backend(
        self, session: WorkbookSession, needed: Literal["file", "xlw"]
    ) -> None:
        """Switch an already-open session's backend in place (PRD sec 6.2.2).

        Save-then-close-then-reopen, strictly in that order, with nothing else interleaved —
        both backends' save()/close() calls are synchronous, so this ordering is what avoids a
        Windows file-lock race (the new backend opening the same scratch path before the old
        one has actually released it).

        Args:
            session: The session to switch — mutated in place (`WorkbookSession` is
                deliberately not frozen).
            needed: The backend to switch to.
        """
        if session.backend == needed:
            return
        logger.info(
            'Switching workbook "%s" backend: %s -> %s',
            session.name,
            session.backend,
            needed,
        )
        if session.dirty:
            if session.backend == "file":
                backends.save_workbook(session.handle, session.path)
            else:
                backends.xlw_save_workbook(session.handle)
            session.dirty = False
        if session.backend == "file":
            backends.close_workbook(session.handle)
        else:
            backends.xlw_close_workbook(session.handle)
        session.handle = self._open_handle(Path(session.path), session.mode, needed)
        session.backend = needed

    def _wire_r4_links_touching(self, name: str) -> None:
        """After staging `name`, repoint (`ChangeLink`) any R4 link between it and another
        workbook that's already staged too, in either direction (plan doc sec 2 R4.1) — a
        no-op until both sides of a given pair have been staged; whichever one gets staged
        second is what actually triggers the repoint.

        Args:
            name: The workbook logical name that was just staged/opened.
        """
        for target in self._link_targets.get(name, ()):
            self._wire_one_r4_link(name, target)
        for source in self._link_sources.get(name, ()):
            self._wire_one_r4_link(source, name)

    def _wire_one_r4_link(self, source: str, target: str) -> None:
        """Repoint `source`'s R4 link(s) to `target` from `target`'s real path to its scratch
        path, once, the first time both are staged (plan doc sec 2 R4.1). Safe even if
        `target`'s scratch copy is a fresh, unedited copy — that's exactly the state its real
        file was already in.

        Args:
            source: The workbook whose link is being repointed. Its session is temporarily
                switched to the `xlw` backend if it isn't already — `ChangeLink` requires COM.
            target: The workbook the link points to, already staged.
        """
        if source not in self._sessions or target not in self._sessions:
            return
        pair = (source, target)
        if pair in self._wired_r4_links:
            return
        self._wired_r4_links.add(pair)
        logger.info('Wiring R4 link: "%s" -> "%s" (staging)', source, target)
        source_session = self._sessions[source]
        if source_session.backend != "xlw":
            self._switch_backend(source_session, "xlw")
        target_real = self._scratch.real_path(target).resolve()
        target_working = self._scratch.working_path(target)
        for current_source_name in backends.com_link_sources(source_session.handle):
            if (
                resolve_link_target(current_source_name, Path(source_session.path))
                == target_real
            ):
                logger.debug(
                    'Repointing link "%s" in "%s" to scratch copy: %s',
                    current_source_name,
                    source,
                    target_working,
                )
                backends.com_change_link(
                    source_session.handle, current_source_name, str(target_working)
                )
                source_session.dirty = True

    def _revert_r4_links_before_commit(self, name: str) -> None:
        """Commit-time hook (plan doc sec 3.2.1): repoint every outbound R4 link `name` has,
        from its target's scratch path back to the target's real path, and save — called by
        `ScratchManager.commit_all()` right before `name`'s own commit, once its target(s) are
        already committed (guaranteed by `commit_order`, R5). No extra recalculation needed
        after this save (probe10).

        Args:
            name: The workbook about to be committed.
        """
        targets = self._link_targets.get(name)
        if not targets:
            return
        session = self._sessions[name]
        if session.backend != "xlw":
            self._switch_backend(session, "xlw")
        for target in targets:
            target_real = self._scratch.real_path(target).resolve()
            target_working = self._scratch.working_path(target).resolve()
            for current_source_name in backends.com_link_sources(session.handle):
                if (
                    resolve_link_target(current_source_name, Path(session.path))
                    == target_working
                ):
                    logger.info(
                        'Reverting R4 link: "%s" -> "%s" (before commit)', name, target
                    )
                    backends.com_change_link(
                        session.handle, current_source_name, str(target_real)
                    )
                    session.dirty = True
        if session.dirty:
            backends.xlw_save_workbook(session.handle)
            session.dirty = False

    def _create(self, ref: WorkbookRef, at_path: Path) -> None:
        if not ref.create_if_missing:
            raise ActionExecutionError(
                ErrorDetail(
                    message=f'Workbook file "{ref.file}" does not exist, and `create_if_missing` is not set.',
                    technical_reason=f"SessionManager: missing file {ref.file!r}, create_if_missing=False",
                )
            )
        template_path = self._workbooks[ref.template].file if ref.template else None
        at_path.parent.mkdir(parents=True, exist_ok=True)
        backends.create_workbook(str(at_path), template_path=template_path)

    def _save_dirty_staged_sessions(self) -> None:
        for session in self._sessions.values():
            if session.scratch_path is not None and session.dirty:
                if session.backend == "file":
                    backends.save_workbook(session.handle, session.path)
                else:
                    backends.xlw_save_workbook(session.handle)
                session.dirty = False

    def checkpoint(self) -> None:
        """Persist every dirty staged session's in-memory state to its scratch file.

        openpyxl writes stay in memory until an explicit save — nothing else flushes them to
        disk mid-run — so without this, a crash after a successful step would leave the
        scratch copy (the recovery artifact, PRD sec 6.3.1) no more informative than the
        original file. `runner.py` calls this after every step, not just at the end (Spec
        sec 6.1) — found necessary via a failing crash-safety integration test, not designed
        up front.
        """
        self._save_dirty_staged_sessions()

    def commit_all(self) -> None:
        """Save every staged session's in-memory state, then commit scratch copies to their
        real paths (PRD sec 6.3.1). Read-only sessions were never staged and aren't touched.
        Redundant with per-step `checkpoint()` calls in the normal case (both are dirty-gated,
        so this is mostly a no-op by the time a run reaches here) — kept anyway as the final
        safety net at the commit boundary, not something to remove just because it's usually
        a no-op.

        Commits in `self._commit_order` (R5) when there's an R4 link graph, so a workbook is
        always committed after every workbook its own R4 link(s) point to; `_revert_r4_links_
        before_commit` runs right before each one's own commit (plan doc sec 3.2).
        """
        self._save_dirty_staged_sessions()
        self._scratch.commit_all(
            order=self._commit_order, before_commit=self._revert_r4_links_before_commit
        )

    def close_all(self) -> None:
        """Close every open session, then quit the shared owned Excel instance (if one was ever
        spawned), attempting all of it even if some steps fail.

        Raises:
            ExceptionGroup: If one or more sessions, or the owned Excel instance, failed to
                close. Every session still gets a close attempt regardless (PRD sec 6.3's
                crash-safety requirement) — this isn't a defensive catch-and-ignore, every
                failure is still surfaced, just after giving every other session (and the
                owned instance) a chance to close too.
        """
        errors: list[Exception] = []
        for session in self._sessions.values():
            try:
                if session.backend == "file":
                    backends.close_workbook(session.handle)
                else:
                    backends.xlw_close_workbook(session.handle)
            except Exception as exc:  # noqa: BLE001 - intentional, see docstring
                errors.append(exc)
        try:
            self._owned_instances.close_owned()
        except ExceptionGroup as exc:
            errors.extend(exc.exceptions)
        if errors:
            raise ExceptionGroup(
                "failed to close one or more workbook sessions", errors
            )


# --- Validation, tier 1: static schema (Spec sec 5.4) --------------------------------------
#
# No workbook access — PRD sec 9.1's fourth example message (checking a range against a
# workbook's *actual* defined names) needs one, which contradicts that constraint. That check
# is not implemented here; it's a documented gap, not a silent omission — see
# docs/Specification.md sec 5.4 for the correction. What's implemented below only needs the
# parsed Workflow structure and the action registry, nothing else.

_IMPLICIT_FIELDS = {
    "workbook"
}  # consumed by the (not yet built) runner before dispatch, not
# part of any action's own param_schema — see Spec sec 4.
_SCHEMA_EXEMPT_ACTIONS = {
    "copy",
    "stop",
    "dump",
}  # copy's raw YAML shape (source/target dicts) doesn't
# match its Python signature yet — needs the runner's
# translation layer (Spec sec 4/8 item 7). stop has no
# workbook: field at all (PRD sec 6.9), so it's exempt from
# the implicit workbook requirement every other action gets.
_STEP_REF_RE = re.compile(r"steps\.([A-Za-z_][A-Za-z0-9_]*)")


def _step_label(step: Step) -> str:
    return f'Step "{step.id}" (action: "{step.action}")'


def _matches_type(value: Any, expected: Any) -> bool:
    """Whether `value` is compatible with a param's declared type annotation.

    Handles plain types, `X | Y` unions (checks any branch matches), `Literal[...]` (checks
    membership), and generic aliases like `list[...]`/`dict[...]` (checks only the origin
    type — e.g. "is this a list at all", not "is every element the right type").
    """
    if expected is inspect.Parameter.empty or expected is Any:
        return True
    origin = typing.get_origin(expected)
    if origin is typing.Literal:
        return value in typing.get_args(expected)
    if origin is pytypes.UnionType or origin is typing.Union:
        return any(_matches_type(value, arg) for arg in typing.get_args(expected))
    if origin is not None:
        return isinstance(value, origin)
    if expected is float:
        # PEP 484's numeric tower: an int is valid anywhere a float is expected (e.g.
        # `set_column_width`'s `width: 20` in YAML, which parses as int, not `20.0`) — bool is
        # technically an int subclass too, but not a sane width/etc. value, so excluded.
        return isinstance(value, float) or (
            isinstance(value, int) and not isinstance(value, bool)
        )
    if isinstance(expected, type):
        return isinstance(value, expected)
    # No current action's signature has an annotation shape that reaches here (every real one
    # is a plain type, a Union, a Literal, or a generic alias — all handled above). Kept as an
    # explicit "don't block on an unrecognized annotation" fallback rather than removed, since
    # mypy needs a return on every path and a future action's signature might reach it.
    return True  # pragma: no cover


def _type_name(expected: Any) -> str:
    origin = typing.get_origin(expected)
    if origin is pytypes.UnionType or origin is typing.Union:
        return " or ".join(
            _type_name(arg)
            for arg in typing.get_args(expected)
            if arg is not type(None)
        )
    if origin is typing.Literal:
        return " or ".join(repr(arg) for arg in typing.get_args(expected))
    if origin is not None:
        return getattr(origin, "__name__", str(origin))
    return getattr(expected, "__name__", str(expected))


def _expects_list(expected: Any) -> bool:
    origin = typing.get_origin(expected)
    if origin is pytypes.UnionType or origin is typing.Union:
        return any(_expects_list(arg) for arg in typing.get_args(expected))
    return origin is list or expected is list


def _type_mismatch_detail(
    step: Step, field: str, value: Any, expected: Any
) -> ErrorDetail:
    suggestion = (
        f"Wrap it in [ ], e.g. [{value}]."
        if _expects_list(expected) and isinstance(value, str)
        else None
    )
    return ErrorDetail(
        message=(
            f'{_step_label(step)}: field "{field}" must be a {_type_name(expected)}, '
            f"got {type(value).__name__} ({value!r})."
        ),
        technical_reason=f"type mismatch: field {field!r} expected {expected!r}, got {type(value)!r}",
        field=field,
        suggestion=suggestion,
    )


def _find_step_refs(value: Any) -> set[str]:
    """Recursively find every `steps.<id>` reference in a raw (unresolved) param value."""
    if isinstance(value, str):
        return set(_STEP_REF_RE.findall(value))
    if isinstance(value, dict):
        refs: set[str] = set()
        for key, val in value.items():
            refs |= _find_step_refs(key)
            refs |= _find_step_refs(val)
        return refs
    if isinstance(value, list):
        refs = set()
        for item in value:
            refs |= _find_step_refs(item)
        return refs
    return set()


def _check_action_exists(
    workflow: Workflow, registry: dict[str, ActionSpec]
) -> ValidationError | None:
    for step in workflow.steps:
        if step.action not in registry:
            suggestion = difflib.get_close_matches(
                step.action, registry.keys(), n=1, cutoff=0.5
            )
            message = f'Step "{step.id}": unknown action "{step.action}".'
            if suggestion:
                message += f' Did you mean "{suggestion[0]}"?'
            return ValidationError(
                ErrorDetail(
                    message=message,
                    technical_reason=f"action {step.action!r} not found in registry",
                    field="action",
                )
            )
    return None


def _check_unknown_params(
    workflow: Workflow, registry: dict[str, ActionSpec]
) -> ValidationError | None:
    for step in workflow.steps:
        if step.action in _SCHEMA_EXEMPT_ACTIONS:
            continue
        allowed = (
            set(registry[step.action].param_schema["properties"]) | _IMPLICIT_FIELDS
        )
        extra = sorted(set(step.params) - allowed)
        if extra:
            return ValidationError(
                ErrorDetail(
                    message=f'{_step_label(step)}: field "{extra[0]}" is not a recognized parameter for this action.',
                    technical_reason=f"unexpected param {extra[0]!r} for action {step.action!r}",
                    field=extra[0],
                )
            )
    return None


def _check_required_params(
    workflow: Workflow, registry: dict[str, ActionSpec]
) -> ValidationError | None:
    for step in workflow.steps:
        if step.action in _SCHEMA_EXEMPT_ACTIONS:
            continue
        required = (
            set(registry[step.action].param_schema["required"]) | _IMPLICIT_FIELDS
        )
        missing = sorted(required - set(step.params))
        if missing:
            return ValidationError(
                ErrorDetail(
                    message=f'{_step_label(step)}: missing required field "{missing[0]}".',
                    technical_reason=f"required param {missing[0]!r} missing for action {step.action!r}",
                    field=missing[0],
                    suggestion=f'Add a "{missing[0]}:" field to this step.',
                )
            )
    return None


def _check_param_types(
    workflow: Workflow, registry: dict[str, ActionSpec]
) -> ValidationError | None:
    for step in workflow.steps:
        if step.action in _SCHEMA_EXEMPT_ACTIONS:
            continue
        properties = registry[step.action].param_schema["properties"]
        for name, value in step.params.items():
            if isinstance(value, str) and is_whole_template_expression(value):
                continue  # can't know its real type until execution — see docstring above
            if name in _IMPLICIT_FIELDS:
                expected: Any = str
            elif name in properties:
                expected = properties[name]["type"]
            else:
                # Unreachable via validate_static's fixed check order (_check_unknown_params
                # always runs first and would already have raised) — kept for when this
                # function is called directly/in isolation, per Spec sec 5.4's "each check
                # individually testable" intent.
                continue  # pragma: no cover
            if not _matches_type(value, expected):
                return ValidationError(
                    _type_mismatch_detail(step, name, value, expected)
                )
    return None


def _check_step_references(
    workflow: Workflow, registry: dict[str, ActionSpec]
) -> ValidationError | None:
    step_index = {step.id: i for i, step in enumerate(workflow.steps)}
    for i, step in enumerate(workflow.steps):
        refs = _find_step_refs(step.params) | (
            _find_step_refs(step.if_expr) if step.if_expr else set()
        )
        for ref in sorted(refs):
            if ref not in step_index:
                suggestion = difflib.get_close_matches(
                    ref, list(step_index), n=1, cutoff=0.5
                )
                message = f'{_step_label(step)}: references step id "{ref}", which does not exist.'
                if suggestion:
                    message += f' Did you mean "{suggestion[0]}" (defined at step {step_index[suggestion[0]] + 1})?'
                return ValidationError(
                    ErrorDetail(
                        message=message,
                        technical_reason=f"unknown step reference {ref!r}",
                    )
                )
            if step_index[ref] >= i:
                return ValidationError(
                    ErrorDetail(
                        message=(
                            f'{_step_label(step)}: references step id "{ref}", which is defined at or after '
                            "this step. A step can only reference an earlier step's output."
                        ),
                        technical_reason=f"forward/self reference to step {ref!r}",
                    )
                )
    return None


_STATIC_CHECKS: list[
    Callable[[Workflow, dict[str, ActionSpec]], ValidationError | None]
] = [
    _check_action_exists,
    _check_unknown_params,
    _check_required_params,
    _check_param_types,
    _check_step_references,
]


def validate_static(workflow: Workflow, registry: dict[str, ActionSpec]) -> None:
    """Tier-1 validation: structural checks against the raw parsed Workflow, no workbook access.

    Args:
        workflow: The parsed workflow to validate.
        registry: The action registry (from `discover_actions`) to check against.

    Raises:
        ValidationError: On the first problem found, in check order — action existence,
            unrecognized params, missing required params, param type mismatches, then
            step-id references (existence and ordering).
    """
    for check in _STATIC_CHECKS:
        error = check(workflow, registry)
        if error is not None:
            raise error


# --- Validation, tier 2: dry-run / step-graph (Spec sec 5.4) --------------------------------


@dataclass(frozen=True)
class ExecutionPlan:
    """Per-workbook mode, inferred from a static pass over the step list (PRD sec 6.3).

    Args:
        modes: Workbook name to "read_only" or "read_write" — read_write iff some step's
            action writes to it (or, for `copy`, iff it's referenced at all — see `plan()`).
    """

    modes: dict[str, Literal["read_only", "read_write"]]


def _find_workbook_names(value: Any) -> set[str]:
    """Recursively find every value under a key literally named "workbook" in a raw param
    structure — handles both a flat `workbook:` field and copy's nested `source`/`target`
    dicts with the same generic walk.
    """
    names: set[str] = set()
    if isinstance(value, dict):
        for key, val in value.items():
            if key == "workbook" and isinstance(val, str):
                names.add(val)
            else:
                names |= _find_workbook_names(val)
    elif isinstance(value, list):
        for item in value:
            names |= _find_workbook_names(item)
    return names


def _check_workbooks_declared(workflow: Workflow) -> ValidationError | None:
    for step in workflow.steps:
        for name in sorted(_find_workbook_names(step.params)):
            if name not in workflow.workbooks:
                return ValidationError(
                    ErrorDetail(
                        message=(
                            f'{_step_label(step)}: references workbook "{name}", which is not declared '
                            "in the workbooks: registry."
                        ),
                        technical_reason=f"undeclared workbook {name!r}",
                    )
                )
    return None


def plan(workflow: Workflow, registry: dict[str, ActionSpec]) -> ExecutionPlan:
    """Tier-2 validation: reasons over the whole step list together, still no workbook access.

    Args:
        workflow: The parsed workflow to plan.
        registry: The action registry (Spec sec 5.1) — consulted for each step's `writes` flag.

    Returns:
        An `ExecutionPlan` with an inferred read/write mode per declared workbook.

    Raises:
        ValidationError: If a step references a workbook not in the `workbooks:` registry.
    """
    error = _check_workbooks_declared(workflow)
    if error is not None:
        raise error
    modes: dict[str, Literal["read_only", "read_write"]] = dict.fromkeys(
        workflow.workbooks, "read_only"
    )
    for step in workflow.steps:
        names = _find_workbook_names(step.params)
        if step.action == "copy":
            # Can't statically tell copy's source from its target without the runner's
            # translation layer (Spec sec 4) — the safe fallback (PRD sec 6.3) is read_write
            # for every workbook it touches, rather than silently under-provisioning one.
            for name in names:
                modes[name] = "read_write"
        elif step.action in registry and registry[step.action].writes:
            for name in names:
                modes[name] = "read_write"
    return ExecutionPlan(modes=modes)


# --- Validation, tier 3: existence check (opt-in, CLI `--check-existence`) -------------------
#
# Unlike tiers 1/2, this one does open workbooks — read-only, via openpyxl, before any
# session/scratch machinery — to confirm every sheet and named range a step references by
# literal name actually exists. Opt-in (not run by default) since it's the first tier that
# touches real files at all. Deliberately does NOT validate plain A1-style cell/range
# references (e.g. "A1", "A1:D6") — only sheet names and workbook-level defined names.

_A1_RANGE_RE = re.compile(r"^\$?[A-Za-z]{1,3}\$?\d+(:\$?[A-Za-z]{1,3}\$?\d+)?$")

# Actions whose `sheet` param must exist (tracked live as create_sheet/rename_sheet/
# delete_sheet steps run earlier in the same workflow).
_SHEET_PARAM_ACTIONS = {
    "read_range",
    "read_metadata",
    "find_headers_row",
    "find_row",
    "find_column",
    "find_columns",
    "write_cell",
    "write_range",
    "write_row",
    "insert_range",
    "set_column_width",
    "rename_sheet",
    "delete_sheet",
}

# Actions whose named field(s) must resolve to a real defined name when the value isn't
# plain A1 notation — nothing in the action set can create a named range, so these are always
# checked against the real workbook's defined_names, never against workflow-tracked state.
_RANGE_PARAM_ACTIONS: dict[str, tuple[str, ...]] = {
    "read_range": ("range",),
    "read_metadata": ("cells",),
    "find_headers_row": ("search_range",),
}


def _looks_like_a1(value: str) -> bool:
    return bool(_A1_RANGE_RE.match(value))


def _sheet_candidates(value: Any) -> list[str]:
    """Extract the literal, checkable sheet name(s) from a `sheet`-shaped param value.

    Handles `read_range`'s multi-sheet forms: a plain string (the common case for every other
    action too), an explicit list, `"all"` (dynamic, nothing to check), or `{"matching": ...}`
    (dynamic, nothing to check) \u2014 same convention as `backends.resolve_sheet_names`. A
    `{{ steps.x... }}` template expression can't be known statically, so it's skipped too.
    """
    if isinstance(value, str):
        if value == "all" or is_whole_template_expression(value):
            return []
        return [value]
    if isinstance(value, list):
        return [
            item
            for item in value
            if isinstance(item, str) and not is_whole_template_expression(item)
        ]
    return []  # dict ("matching") or None \u2014 dynamic/unspecified, nothing to check


def _resolve_check_path(name: str, workflow: Workflow) -> Path | None:
    """The real file to open read-only for `name`'s existence checks, or None if there isn't
    one yet (a fresh `create_if_missing` workbook with no template \u2014 nothing to check).
    """
    ref = workflow.workbooks[name]
    direct = Path(ref.file)
    if direct.exists():
        return direct
    if ref.template is not None:
        template_path = Path(workflow.workbooks[ref.template].file)
        if template_path.exists():
            return template_path
    return None


def _sheet_error(
    step: Step, workbook_name: str, sheet: str, known: set[str]
) -> ValidationError:
    suggestion = difflib.get_close_matches(sheet, sorted(known), n=1, cutoff=0.5)
    message = (
        f'{_step_label(step)}: sheet "{sheet}" does not exist in workbook "{workbook_name}" '
        "and is not created by an earlier step."
    )
    if suggestion:
        message += f' Did you mean "{suggestion[0]}"?'
    return ValidationError(
        ErrorDetail(
            message=message, technical_reason=f"unknown sheet {sheet!r}", field="sheet"
        )
    )


def _range_error(
    step: Step, workbook_name: str, field: str, value: str
) -> ValidationError:
    return ValidationError(
        ErrorDetail(
            message=(
                f'{_step_label(step)}: "{value}" is not a defined name in workbook '
                f'"{workbook_name}".'
            ),
            technical_reason=f"unknown defined name {value!r}",
            field=field,
        )
    )


def _check_copy_existence(step: Step, known_sheets: dict[str, set[str]]) -> None:
    for side in ("source", "target"):
        ref = step.params.get(side)
        if not isinstance(ref, dict):
            continue
        wb_name: str | None = ref.get("workbook")
        for sheet in _sheet_candidates(ref.get("sheet")):
            if wb_name in known_sheets and sheet not in known_sheets[wb_name]:
                raise _sheet_error(step, wb_name, sheet, known_sheets[wb_name])


def _check_create_sheet_existence(
    step: Step, wb_name: str | None, known_sheets: dict[str, set[str]]
) -> None:
    new_name = step.params.get("name")
    if (
        wb_name in known_sheets
        and isinstance(new_name, str)
        and not is_whole_template_expression(new_name)
    ):
        known_sheets[wb_name].add(new_name)


def _check_rename_sheet_existence(
    step: Step, wb_name: str | None, known_sheets: dict[str, set[str]]
) -> None:
    old_name = step.params.get("sheet")
    new_name = step.params.get("new_name")
    if wb_name in known_sheets and isinstance(old_name, str):
        if old_name not in known_sheets[wb_name]:
            raise _sheet_error(step, wb_name, old_name, known_sheets[wb_name])
        if isinstance(new_name, str) and not is_whole_template_expression(new_name):
            known_sheets[wb_name].discard(old_name)
            known_sheets[wb_name].add(new_name)


def _check_delete_sheet_existence(
    step: Step, wb_name: str | None, known_sheets: dict[str, set[str]]
) -> None:
    sheet_to_delete = step.params.get("sheet")
    if wb_name in known_sheets and isinstance(sheet_to_delete, str):
        if sheet_to_delete not in known_sheets[wb_name]:
            raise _sheet_error(step, wb_name, sheet_to_delete, known_sheets[wb_name])
        known_sheets[wb_name].discard(sheet_to_delete)


def _check_sheet_field_existence(
    step: Step, wb_name: str, known_sheets: dict[str, set[str]]
) -> None:
    for sheet in _sheet_candidates(step.params.get("sheet")):
        if sheet not in known_sheets[wb_name]:
            raise _sheet_error(step, wb_name, sheet, known_sheets[wb_name])


def _check_range_field_existence(
    step: Step, wb_name: str, defined_names: dict[str, set[str]]
) -> None:
    for field in _RANGE_PARAM_ACTIONS.get(step.action, ()):
        value = step.params.get(field)
        candidates = value if isinstance(value, list) else [value]
        for candidate in candidates:
            if not isinstance(candidate, str) or is_whole_template_expression(
                candidate
            ):
                continue
            if _looks_like_a1(candidate):
                continue
            if candidate not in defined_names[wb_name]:
                raise _range_error(step, wb_name, field, candidate)


_STRUCTURAL_ACTION_HANDLERS = {
    "create_sheet": _check_create_sheet_existence,
    "rename_sheet": _check_rename_sheet_existence,
    "delete_sheet": _check_delete_sheet_existence,
}


def _check_step_existence(
    step: Step,
    known_sheets: dict[str, set[str]],
    defined_names: dict[str, set[str]],
) -> None:
    if step.action == "copy":
        _check_copy_existence(step, known_sheets)
        return

    wb_name = step.params.get("workbook")

    handler = _STRUCTURAL_ACTION_HANDLERS.get(step.action)
    if handler is not None:
        handler(step, wb_name, known_sheets)
        return

    if wb_name not in known_sheets:
        return  # workbook not checkable (doesn't exist yet) or undeclared (tier 2 already caught)

    if step.action in _SHEET_PARAM_ACTIONS:
        _check_sheet_field_existence(step, wb_name, known_sheets)

    if step.action == "recalculate" and step.params.get("scope") == "sheet":
        _check_sheet_field_existence(step, wb_name, known_sheets)

    _check_range_field_existence(step, wb_name, defined_names)


def validate_existence(workflow: Workflow) -> None:
    """Tier-3 validation (opt-in): confirms every sheet and workbook-level defined name a step
    references by literal name actually exists in the real workbook \u2014 read-only, via openpyxl,
    before any `SessionManager`/`ScratchManager` involvement.

    Only literal, non-templated string references are checked; a `{{ steps.x... }}` expression
    can't be known until execution (same bypass tier 1's `_check_param_types` uses). Plain
    A1-style cell/range references (e.g. "A1", "A1:D6") are never checked \u2014 only sheet names
    and defined names. A sheet created earlier in the same workflow by `create_sheet` (or
    renamed/removed by `rename_sheet`/`delete_sheet`) is tracked step by step, in order, so
    it counts as existing/not-existing from that point on, not just at the real file's
    as-loaded state.

    Args:
        workflow: The parsed workflow to validate.

    Raises:
        ValidationError: On the first missing sheet or missing defined name found.
    """
    opened: dict[Path, Any] = {}
    known_sheets: dict[str, set[str]] = {}
    defined_names: dict[str, set[str]] = {}

    try:
        for name in workflow.workbooks:
            path = _resolve_check_path(name, workflow)
            if path is None:
                continue
            if path not in opened:
                opened[path] = openpyxl.load_workbook(path, read_only=True)
            wb = opened[path]
            known_sheets[name] = set(wb.sheetnames)
            defined_names[name] = set(wb.defined_names.keys())

        for step in workflow.steps:
            _check_step_existence(step, known_sheets, defined_names)
    finally:
        for wb in opened.values():
            wb.close()
