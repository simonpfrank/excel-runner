"""Action functions. See docs/Specification.md sec 4 and docs/PRD.md sec 7 for the catalog.

Every action has the shape `fn(session: WorkbookSession, **params) -> ActionResult`. Note
`workbook` is deliberately absent from every signature below, even though it's a required YAML
field on every step (PRD sec 7/sec 11): the (not-yet-built) runner resolves `workbook` into the
`session` it passes in, so the action itself never needs it as a separate parameter — passing
both would just be the same information twice.

Functions are grouped to match the PRD sec 7 catalog order (basic -> data -> ...) via comment
banners, since there's no file boundary to do it now that actions live in one module
(docs/Specification.md sec 4).
"""

import json
import logging
from pathlib import Path
from typing import Any, Literal

from openpyxl.utils import column_index_from_string, get_column_letter

from excel_runner import backends
from excel_runner.core import (
    ActionExecutionError,
    ActionResult,
    ErrorDetail,
    WorkbookSession,
    com_action,
    control_action,
    file_action,
)

logger = logging.getLogger(__name__)

# --- basic -------------------------------------------------------------------------------


@file_action
def open(session: WorkbookSession) -> ActionResult:
    """Confirm a workbook is open.

    The runner resolves and opens `session` before dispatching to any action (Spec sec 6.1) —
    this exists for explicit manual control and audit-log clarity, not to do the opening
    itself. `update_links` and a `mode` override are not yet parameters here: `update_links`
    has no effect without a live Excel session (COM, a later phase per PRD sec 8), and a mode
    override depends on the read/write inference that static validation (Spec sec 5.4) hasn't
    been built yet to override.

    Args:
        session: The already-open workbook session.

    Returns:
        A success result with no meaningful output.
    """
    return ActionResult(status="success", output={})


@file_action(writes=True)
def save(session: WorkbookSession) -> ActionResult:
    """Save the workbook to its current session path.

    Args:
        session: The workbook session to save.

    Returns:
        A success result with no meaningful output.
    """
    backends.save_workbook(session.handle, session.path)
    return ActionResult(status="success", output={})


@file_action
def close(session: WorkbookSession) -> ActionResult:
    """Close the workbook, releasing its file handle.

    Args:
        session: The workbook session to close.

    Returns:
        A success result with no meaningful output.
    """
    backends.close_workbook(session.handle)
    return ActionResult(status="success", output={})


@com_action(writes=True)
def recalculate(
    session: WorkbookSession,
    scope: Literal["sheet", "workbook", "all"] = "workbook",
    mode: Literal["normal", "full", "full_rebuild"] = "normal",
    sheet: str | None = None,
) -> ActionResult:
    """Force Excel to recalculate formulas in a live session, then save the result immediately.

    Requires a live Excel session (the `com` capability puts `session` on the `xlw` backend,
    switching it there automatically if it wasn't already — PRD sec 6.2.2). Always saves
    before returning, regardless of the session's dirty-tracking, so the recalculated values
    are on disk immediately rather than deferred to end-of-run commit.

    `mode: "full"`/`"full_rebuild"` are always application-wide in Excel — there is no
    per-workbook or per-sheet equivalent in the COM object model — so they require
    `scope: "all"`.

    Args:
        session: The workbook session to recalculate (switched to the `xlw` backend first if
            needed).
        scope: What to recalculate — `"sheet"` (one worksheet), `"workbook"` (default, every
            sheet in this workbook), or `"all"` (every workbook open in this run's shared
            Excel instance).
        mode: `"normal"` (default, only cells Excel considers dirty), `"full"` (force every
            formula to recompute), or `"full_rebuild"` (force recompute and re-check
            dependency trees too). `"full"`/`"full_rebuild"` require `scope: "all"`.
        sheet: Worksheet name, only meaningful when `scope: "sheet"`. If omitted, the
            workbook's active sheet is used instead, and the result's `output.warning` names
            which sheet that was.

    Returns:
        A success result. `output` echoes the effective `scope`/`mode`, plus `sheet` when
        `scope` is `"sheet"`, plus `warning` if `sheet` had to fall back to the active sheet.

    Raises:
        ActionExecutionError: If `sheet` is given with a `scope` other than `"sheet"`
            (ambiguous), or if `mode` is `"full"`/`"full_rebuild"` with a `scope` other than
            `"all"` (not possible in the Excel object model).
    """
    if sheet is not None and scope != "sheet":
        raise ActionExecutionError(
            ErrorDetail(
                message=(
                    f'recalculate: `sheet` was given but `scope` is "{scope}" — `sheet` only '
                    'applies when `scope` is "sheet".'
                ),
                technical_reason=f"recalculate called with sheet={sheet!r}, scope={scope!r}",
            )
        )
    if mode in ("full", "full_rebuild") and scope != "all":
        raise ActionExecutionError(
            ErrorDetail(
                message=(
                    f'recalculate: mode "{mode}" is always application-wide in Excel (there is '
                    f'no per-{scope} equivalent) — use scope: "all", or mode: "normal" for a '
                    "single sheet/workbook."
                ),
                technical_reason=f"recalculate called with mode={mode!r}, scope={scope!r}",
            )
        )

    book = session.handle
    app = book.app
    output: dict[str, Any] = {"scope": scope, "mode": mode}

    if scope == "sheet":
        if sheet is None:
            sheet = book.sheets.active.name
            output["warning"] = (
                f'recalculate: `sheet` not specified — used the active sheet "{sheet}".'
            )
        assert sheet is not None
        output["sheet"] = sheet
        backends.com_calculate_sheet(book, sheet)
    elif scope == "workbook":
        backends.com_calculate_workbook(book)
    elif mode == "normal":
        backends.xlw_calculate_all(app)
    elif mode == "full":
        backends.com_calculate_full(app)
    else:
        backends.com_calculate_full_rebuild(app)

    backends.com_wait_until_calculation_done(app)
    backends.xlw_save_workbook(book)
    session.dirty = False

    return ActionResult(status="success", output=output)


@control_action
def stop(reason: str | None = None) -> ActionResult:
    """Halt the run — no later step runs (PRD sec 6.9).

    No `session` parameter and no `workbook:` field — this is pure control flow, not a backend
    call. Reaching this action is not itself a failure; `runner.py` is what actually ends the
    loop once this returns, marking every later step `"stopped"` (Spec sec 6.1).

    Args:
        reason: Optional note for the audit log explaining why the run stopped.

    Returns:
        A success result; `reason` is echoed into `output` when given, for the audit log.
    """
    return ActionResult(
        status="success", output={"reason": reason} if reason is not None else {}
    )


@control_action
def dump(
    step_outputs: dict[str, dict[str, Any]],
    ids: list[str] | None = None,
    to: Literal["console", "file"] = "console",
    path: str | None = None,
) -> ActionResult:
    """Print (or write) a formatted JSON snapshot of steps' recorded output — the same
    `steps.<id>.output` data `{{ }}` templating reads from — for seeing a workflow's internal
    state while authoring/debugging. No `session`/`workbook:` field — pure introspection, no
    backend call, never mutates a workbook.

    `step_outputs` is injected by the runner (every step's `{"status": ..., "output": ...}` so
    far, in execution order) — never a YAML field, excluded from this action's param schema
    the same way `session` is excluded from every other action's.

    Args:
        step_outputs: Every step's recorded output so far.
        ids: Step ids to include. Omit to include every step that has run so far. An id with no
            recorded output (typo, or genuinely hasn't run yet) is logged as a warning and
            skipped, not raised — a debugging aid shouldn't halt a run over an unknown id.
        to: "console" (default) prints to stdout; "file" writes to `path` instead.
        path: Required when `to` is "file" — where to write the JSON.

    Returns:
        A success result with no meaningful output.

    Raises:
        ActionExecutionError: If `to` is "file" but `path` wasn't given.
    """
    if ids is None:
        selected = step_outputs
    else:
        selected = {}
        for step_id in ids:
            if step_id not in step_outputs:
                logger.warning(
                    'dump: step id "%s" has no recorded output (typo, or has not run yet) — '
                    "skipping.",
                    step_id,
                )
                continue
            selected[step_id] = step_outputs[step_id]

    payload = json.dumps(selected, indent=2, default=str)
    if to == "file":
        if path is None:
            raise ActionExecutionError(
                ErrorDetail(
                    message='dump: to: "file" requires a `path:`.',
                    technical_reason="dump called with to=file but path=None",
                )
            )
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        Path(path).write_text(payload)
    else:
        print(f"--- dump ({len(selected)} step(s)) ---\n{payload}")
    return ActionResult(status="success", output={})


# --- data ----------------------------------------------------------------------------------


def _open_for_formula_read(session: WorkbookSession) -> Any:
    """Open a throwaway `data_only=False` view of the session's current path for a
    `formula: true` request, saving any pending writes first so the on-disk view is current.

    `data_only` is a load-time decision (backends.open_workbook's docstring) — it cannot be
    toggled on session.handle itself, hence the one-off reopen rather than reusing it.

    Args:
        session: The workbook session `formula: true` was requested against.

    Returns:
        A fresh openpyxl Workbook, read-only and data_only=False. Caller must close it.
    """
    if session.dirty:
        backends.save_workbook(session.handle, session.path)
        session.dirty = False
    return backends.open_workbook_for_formula_read(session.path)


@com_action(writes=True)
def copy(
    session: WorkbookSession,
    target: WorkbookSession,
    source_sheet: str,
    target_sheet: str,
    target_range: str,
    source_range: str | None = None,
) -> ActionResult:
    """Copy a range — or, if `source_range` is omitted, the whole sheet — into another session,
    via Excel's own Copy (COM) so formulas and formatting come across too, not just values.

    The one action needing two open sessions at once. The (not-yet-built) runner will need
    special-case wiring to resolve both `source.workbook` and `target.workbook` into `session`
    and `target` before calling this — every other action's single `workbook:` field maps to
    one `session` param, but copy's YAML shape has two nested workbook refs (PRD sec 7/sec 11).

    Args:
        session: The source workbook session (switched to the `xlw` backend first if needed,
            same as every other `com` capability action — PRD sec 6.2.2).
        target: The target workbook session (also switched to the `xlw` backend first).
        source_sheet: Source worksheet name.
        target_sheet: Target worksheet name.
        target_range: Where to start writing — only the top-left cell is used.
        source_range: An A1-style range, or None to copy the whole used range of the sheet.

    Returns:
        A success result with no meaningful output.
    """
    backends.com_copy_range(
        session.handle,
        source_sheet,
        source_range,
        target.handle,
        target_sheet,
        target_range,
    )
    target.dirty = True
    return ActionResult(status="success", output={})


@file_action
def read_range(
    session: WorkbookSession,
    sheet: str | list[str] | dict[str, str],
    range: str,
    formula: bool = False,
) -> ActionResult:
    """Read a cell or range of cells, from one sheet or several.

    Defaults to each cell's computed value (the file-backend session opens with
    `data_only=True`, backends.open_workbook's docstring); pass `formula: true` to read the
    formula text instead, via a one-off reopen (`_open_for_formula_read`) — the session's main
    handle is never touched.

    Args:
        session: The workbook session to read from.
        sheet: A single worksheet name (plain string), an explicit list of names (multi-sheet
            capture), the literal string `"all"` (every sheet in the workbook), or
            `{"matching": <regex>}` (every sheet whose name matches, via `re.search` — same
            convention as `find_row`/`find_headers_row`'s `patterns`). See PRD sec 7.
        range: An A1-style cell (e.g. "B2") or range (e.g. "A1:D50"), or a workbook-level
            defined name — same for every sheet read.
        formula: If True, read formula text instead of each cell's computed value.

    Returns:
        `{"values": ...}` (PRD sec 10.4's output-shape rule: always a keyed object). For a
        single sheet name, `values` is that sheet's cell value or 2D list, unchanged from
        before this was multi-sheet-aware. For a list/`"all"`/`matching`, `values` is a dict
        keyed by sheet name, one entry per resolved sheet.

    Raises:
        ActionExecutionError: If `range` is neither valid A1 notation nor a real defined name
            in the workbook, or is a defined name spanning more than one area.
    """
    workbook = _open_for_formula_read(session) if formula else session.handle
    try:
        try:
            if isinstance(sheet, str) and sheet != "all":
                values = backends.read_range(workbook, sheet, range)
                return ActionResult(status="success", output={"values": values})
            sheet_names = backends.resolve_sheet_names(workbook, sheet)
            values_by_sheet = {
                name: backends.read_range(workbook, name, range) for name in sheet_names
            }
            return ActionResult(status="success", output={"values": values_by_sheet})
        except ValueError as exc:
            raise ActionExecutionError(
                ErrorDetail(
                    message=f'read_range: "{range}" is not a valid range or defined name.',
                    technical_reason=f"{type(exc).__name__}: {exc}",
                )
            ) from exc
    finally:
        if formula:
            backends.close_workbook(workbook)


@file_action
def read_metadata(
    session: WorkbookSession,
    target: Literal["properties", "cells"],
    sheet: str | None = None,
    cells: list[str] | None = None,
    formula: bool = False,
) -> ActionResult:
    """Read document properties, or a scattered list of specific cells.

    The `textboxes` sub-target from PRD sec 7 is COM-only (openpyxl can't see live control
    state) and not built here — deferred to the COM phase (Spec sec 8). `sheet` is a
    clarification found during implementation: PRD sec 7's catalog didn't list it for the
    `cells` sub-case, but reading specific cells needs to know which worksheet they're on,
    same as every other cell-addressing action.

    Args:
        session: The workbook session to read from.
        target: "properties" for document properties, "cells" for a scattered cell list.
        sheet: Worksheet name — required if target is "cells".
        cells: A1-style cell references to read, or workbook-level defined names — required
            if target is "cells".
        formula: If True and target is "cells", read formula text instead of each cell's
            computed value — same one-off reopen as `read_range`'s `formula:` param.

    Returns:
        `{"values": ...}`-style keyed output: document properties by name, or cell reference
        to value, depending on `target`.

    Raises:
        ActionExecutionError: If target is "cells" but `sheet`/`cells` weren't given, or if a
            cell reference is neither valid A1 notation nor a real defined name in the
            workbook, or is a defined name spanning more than one area.
    """
    if target == "properties":
        return ActionResult(
            status="success", output=backends.read_properties(session.handle)
        )
    if target != "cells":
        # Python doesn't enforce type hints at runtime, and this function is directly
        # importable/callable on its own (PRD sec 3/sec 9's library goal) — so an unsupported
        # target must be rejected explicitly here, not silently fall through to the "cells"
        # handling below just because it wasn't "properties". Found while reasoning about
        # target="textboxes" specifically: it used to be mishandled exactly this way.
        raise ActionExecutionError(
            ErrorDetail(
                message=f'read_metadata: target "{target}" is not supported yet.',
                technical_reason=f"read_metadata called with unsupported target {target!r}",
            )
        )
    if sheet is None or cells is None:
        raise ActionExecutionError(
            ErrorDetail(
                message='read_metadata: target "cells" requires both `sheet` and `cells`.',
                technical_reason="read_metadata called with target=cells but sheet or cells was None",
            )
        )
    workbook = _open_for_formula_read(session) if formula else session.handle
    try:
        try:
            return ActionResult(
                status="success", output=backends.read_cells(workbook, sheet, cells)
            )
        except ValueError as exc:
            raise ActionExecutionError(
                ErrorDetail(
                    message=f"read_metadata: one of {cells} is not a valid cell or defined name.",
                    technical_reason=f"{type(exc).__name__}: {exc}",
                )
            ) from exc
    finally:
        if formula:
            backends.close_workbook(workbook)


@file_action(writes=True)
def write_cell(
    session: WorkbookSession, sheet: str, cell: str, value: Any
) -> ActionResult:
    """Write a value to a single cell.

    Args:
        session: The workbook session to write to.
        sheet: Worksheet name.
        cell: An A1-style cell reference (e.g. "B2").
        value: The value to write. A string starting with "=" is stored as a formula.

    Returns:
        A success result with no meaningful output.
    """
    backends.write_cell(session.handle, sheet, cell, value)
    session.dirty = True
    return ActionResult(status="success", output={})


@file_action(writes=True)
def write_range(
    session: WorkbookSession, sheet: str, range: str, values: list[list[Any]]
) -> ActionResult:
    """Write a 2D block of values, anchored at the top-left cell of `range`.

    Args:
        session: The workbook session to write to.
        sheet: Worksheet name.
        range: An A1-style cell or range — only the top-left cell is used as the anchor.
        values: A 2D list of row values to write.

    Returns:
        A success result with no meaningful output.
    """
    backends.write_range(session.handle, sheet, range, values)
    session.dirty = True
    return ActionResult(status="success", output={})


@file_action(writes=True)
def write_row(
    session: WorkbookSession,
    sheet: str,
    row: int,
    values: dict[str, Any] | list[Any],
    start_column: str | None = None,
) -> ActionResult:
    """Write a row of values, either by explicit column mapping or positionally.

    The by-header mode (`values_by_header` + `headers_from`, PRD sec 7/sec 11 item 9) isn't
    built here — it needs another step's output, which doesn't exist as a concept until
    runner.py threads step-output context through (Spec sec 4/8).

    Args:
        session: The workbook session to write to.
        sheet: Worksheet name.
        row: The row number to write into.
        values: Either `{column: value}` (explicit mapping), or a plain ordered list written
            left-to-right starting at `start_column` (positional mode).
        start_column: Required when `values` is a list — the column to start writing at.

    Returns:
        A success result with no meaningful output.

    Raises:
        ActionExecutionError: If `values` is a list but `start_column` wasn't given.
    """
    if isinstance(values, dict):
        for column, value in values.items():
            backends.write_cell(session.handle, sheet, f"{column}{row}", value)
    else:
        if start_column is None:
            raise ActionExecutionError(
                ErrorDetail(
                    message="write_row: positional `values` (a list) needs `start_column`.",
                    technical_reason="write_row called with list values but start_column=None",
                )
            )
        start_idx = column_index_from_string(start_column)
        for offset, value in enumerate(values):
            column = get_column_letter(start_idx + offset)
            backends.write_cell(session.handle, sheet, f"{column}{row}", value)
    session.dirty = True
    return ActionResult(status="success", output={})


# --- structure -----------------------------------------------------------------------------


@file_action(writes=True)
def insert_range(
    session: WorkbookSession,
    sheet: str,
    at: str,
    direction: Literal["rows", "columns"] | None = None,
    header: dict[str, Any] | None = None,
) -> ActionResult:
    """Insert a whole row or whole column, shifting existing content.

    A partial range (e.g. "C5:C10") isn't built yet (PRD sec 11 item 12's flagged cost) —
    caught here and returned as a structured error, consistent with find_*'s "legitimately
    didn't work" pattern below, rather than a raw exception escaping.

    Args:
        session: The workbook session to modify.
        sheet: Worksheet name.
        at: A whole-column reference (e.g. "C:C") or whole-row reference (e.g. "5:5").
        direction: Unused for whole-row/whole-column inserts (unambiguous from `at` itself).
        header: `{"row": int, "text": str}` — only meaningful for a column insert.

    Returns:
        A success result, or a structured error if `at` is a partial range.
    """
    try:
        backends.insert_range(session.handle, sheet, at, direction, header)
    except NotImplementedError as exc:
        return ActionResult(
            status="error",
            output={},
            error=ErrorDetail(
                message=str(exc), technical_reason=f"{type(exc).__name__}: {exc}"
            ),
        )
    session.dirty = True
    return ActionResult(status="success", output={})


@file_action(writes=True)
def set_column_width(
    session: WorkbookSession,
    sheet: str,
    columns: str,
    width: float | Literal["autofit"],
) -> ActionResult:
    """Set the width of a column or range of columns.

    Args:
        session: The workbook session to modify.
        sheet: Worksheet name.
        columns: A single column letter (e.g. "B") or range (e.g. "A:C").
        width: An explicit width, or "autofit".

    Returns:
        A success result with no meaningful output.
    """
    backends.set_column_width(session.handle, sheet, columns, width)
    session.dirty = True
    return ActionResult(status="success", output={})


@file_action(writes=True)
def create_sheet(
    session: WorkbookSession, name: str, index: int | None = None
) -> ActionResult:
    """Add a new, empty worksheet to the workbook.

    Args:
        session: The workbook session to modify.
        name: Name for the new sheet.
        index: Position to insert at (0-based). Appended at the end if omitted.

    Returns:
        A success result, or a structured error if a sheet named `name` already exists —
        a normal "didn't work" outcome (a workflow author picked a name already in use), not
        an unexpected failure, consistent with the find_*/insert_range error pattern above.
    """
    try:
        backends.create_sheet(session.handle, name, index)
    except ValueError as exc:
        return ActionResult(
            status="error",
            output={},
            error=ErrorDetail(
                message=str(exc), technical_reason=f"{type(exc).__name__}: {exc}"
            ),
        )
    session.dirty = True
    return ActionResult(status="success", output={})


@file_action(writes=True)
def rename_sheet(session: WorkbookSession, sheet: str, new_name: str) -> ActionResult:
    """Rename an existing worksheet.

    Args:
        session: The workbook session to modify.
        sheet: Current worksheet name.
        new_name: New name for the worksheet.

    Returns:
        A success result with no meaningful output.
    """
    backends.rename_sheet(session.handle, sheet, new_name)
    session.dirty = True
    return ActionResult(status="success", output={})


@file_action(writes=True)
def delete_sheet(session: WorkbookSession, sheet: str) -> ActionResult:
    """Remove a worksheet from the workbook.

    Args:
        session: The workbook session to modify.
        sheet: Name of the worksheet to remove.

    Returns:
        A success result, or a structured error if `sheet` is the workbook's only remaining
        sheet — a workbook can't have zero sheets, and this is a normal, avoidable outcome for
        a workflow author to react to, not an unexpected failure.
    """
    try:
        backends.delete_sheet(session.handle, sheet)
    except ValueError as exc:
        return ActionResult(
            status="error",
            output={},
            error=ErrorDetail(
                message=str(exc), technical_reason=f"{type(exc).__name__}: {exc}"
            ),
        )
    session.dirty = True
    return ActionResult(status="success", output={})


# --- lookup ----------------------------------------------------------------------------------


@file_action
def find_headers_row(
    session: WorkbookSession, sheet: str, search_range: str, patterns: list[str]
) -> ActionResult:
    """Find the row within `search_range` where every pattern matches some cell in that row.

    Args:
        session: The workbook session to search.
        sheet: Worksheet name.
        search_range: An A1-style range to search within, or a workbook-level defined name.
        patterns: Regex patterns — every one must match a cell in a row for that row to count.

    Returns:
        `{"row": int, "headers": {pattern: column_letter}}`, or a structured error if no row
        matches every pattern — a normal outcome of a search, not an unexpected failure.

    Raises:
        ActionExecutionError: If `search_range` is neither valid A1 notation nor a real
            defined name in the workbook, or is a defined name spanning more than one area.
    """
    try:
        result = backends.find_headers_row(
            session.handle, sheet, search_range, patterns
        )
    except ValueError as exc:
        raise ActionExecutionError(
            ErrorDetail(
                message=f'find_headers_row: "{search_range}" is not a valid range or defined name.',
                technical_reason=f"{type(exc).__name__}: {exc}",
            )
        ) from exc
    if result is None:
        return ActionResult(
            status="error",
            output={},
            error=ErrorDetail(
                message=f'No row in "{search_range}" of sheet "{sheet}" matches every pattern: {patterns}.',
                technical_reason="find_headers_row: no row matched all patterns",
            ),
        )
    row, headers = result
    return ActionResult(status="success", output={"row": row, "headers": headers})


@file_action
def find_row(
    session: WorkbookSession,
    sheet: str,
    column: str,
    search_value: Any,
    header_row: int | None = None,
) -> ActionResult:
    """Find the row number where `column` equals `search_value`.

    Args:
        session: The workbook session to search.
        sheet: Worksheet name.
        column: A column letter (e.g. "B").
        search_value: The value to match, by equality.
        header_row: If given, search starts on the row after it.

    Returns:
        `{"row": int}`, or a structured error if not found.
    """
    row = backends.find_row(session.handle, sheet, column, search_value, header_row)
    if row is None:
        return ActionResult(
            status="error",
            output={},
            error=ErrorDetail(
                message=f'No row found in column "{column}" of sheet "{sheet}" matching "{search_value}".',
                technical_reason="find_row: no matching row",
            ),
        )
    return ActionResult(status="success", output={"row": row})


@file_action
def find_column(
    session: WorkbookSession, sheet: str, header_row: int, pattern: str
) -> ActionResult:
    """Find the column letter whose header (in `header_row`) matches `pattern`.

    Args:
        session: The workbook session to search.
        sheet: Worksheet name.
        header_row: The row number containing headers.
        pattern: A regex pattern to match against each header cell's value.

    Returns:
        `{"column": str}`, or a structured error if not found.
    """
    column = backends.find_column(session.handle, sheet, header_row, pattern)
    if column is None:
        return ActionResult(
            status="error",
            output={},
            error=ErrorDetail(
                message=f'No column found in sheet "{sheet}" row {header_row} matching "{pattern}".',
                technical_reason="find_column: no matching column",
            ),
        )
    return ActionResult(status="success", output={"column": column})


@file_action
def find_columns(
    session: WorkbookSession, sheet: str, header_row: int, patterns: dict[str, str]
) -> ActionResult:
    """Find multiple named columns by header pattern in one call.

    Args:
        session: The workbook session to search.
        sheet: Worksheet name.
        header_row: The row number containing headers.
        patterns: Logical name to regex pattern.

    Returns:
        Logical name to column letter, for every pattern that matched. Names whose pattern
        didn't match anything are simply absent — not an error at this level (PRD sec 10.4).
    """
    result = backends.find_columns(session.handle, sheet, header_row, patterns)
    return ActionResult(status="success", output=result)
