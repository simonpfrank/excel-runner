"""Integration tests for run_workflow() — the full stack, zero mocks (project convention):
real YAML files, real openpyxl workbooks, no fakes anywhere. This is the first point genuine
end-to-end tests are possible (Spec sec 6.1/6.2, build order item 7) and the shape the user and
Claude agreed integration tests for this project should mostly take: a real workflow.yaml run
through run_workflow(), asserted against the resulting real workbook state.
"""

import json
from pathlib import Path

import openpyxl
import pytest

from excel_runner.core import ActionExecutionError, ValidationError
from excel_runner.runner import run_workflow


def _make_workbook(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet"
    sheet["A1"] = "hello"
    workbook.save(path)
    return path


def _write_yaml(path: Path, text: str) -> Path:
    path.write_text(text)
    return path


class TestHappyPath:
    def test_reads_writes_and_implicitly_saves_on_success(self, tmp_path: Path) -> None:
        """No explicit save/close step — commit_all() at the end of a successful run must
        persist the change to the real file on its own (PRD sec 6.3's implicit save)."""
        real = _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: get_a1
                action: read_range
                workbook: manip
                sheet: "Sheet"
                range: "A1"
              - id: write_b1
                action: write_cell
                workbook: manip
                sheet: "Sheet"
                cell: "B1"
                value: "{{ steps.get_a1.output.values }}"
            """,
        )

        result = run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert result.status == "success"
        reopened = openpyxl.load_workbook(real)
        assert reopened["Sheet"]["B1"].value == "hello"

    def test_explicit_save_and_close_steps_still_work(self, tmp_path: Path) -> None:
        real = _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: write_b1
                action: write_cell
                workbook: manip
                sheet: "Sheet"
                cell: "B1"
                value: "explicit"
              - id: save_it
                action: save
                workbook: manip
              - id: close_it
                action: close
                workbook: manip
            """,
        )

        result = run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert result.status == "success"
        assert openpyxl.load_workbook(real)["Sheet"]["B1"].value == "explicit"


class TestIfConditions:
    def test_skipped_step_does_not_run(self, tmp_path: Path) -> None:
        _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: find_it
                action: find_row
                workbook: manip
                sheet: "Sheet"
                column: "A"
                search_value: "does-not-exist"
              - id: write_if_found
                action: write_cell
                workbook: manip
                sheet: "Sheet"
                cell: "C1"
                value: "should not appear"
                if: "{{ steps.find_it.status == 'success' }}"
            """,
        )

        result = run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert result.status == "error"  # find_it itself failed
        skipped = [s for s in result.step_results if s.step_id == "write_if_found"][0]
        assert skipped.status == "skipped"


class TestStepFailureDoesNotCrashTheRun:
    def test_a_normal_search_miss_does_not_stop_later_steps(self, tmp_path: Path) -> None:
        _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: find_it
                action: find_row
                workbook: manip
                sheet: "Sheet"
                column: "A"
                search_value: "does-not-exist"
              - id: still_runs
                action: read_range
                workbook: manip
                sheet: "Sheet"
                range: "A1"
            """,
        )

        result = run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert result.status == "error"
        statuses = {s.step_id: s.status for s in result.step_results}
        assert statuses["find_it"] == "error"
        assert statuses["still_runs"] == "success"

    def test_a_failed_run_never_commits_to_the_real_file(self, tmp_path: Path) -> None:
        real = _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: write_b1
                action: write_cell
                workbook: manip
                sheet: "Sheet"
                cell: "B1"
                value: "should not be committed"
              - id: find_it
                action: find_row
                workbook: manip
                sheet: "Sheet"
                column: "A"
                search_value: "does-not-exist"
            """,
        )

        result = run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert result.status == "error"
        assert openpyxl.load_workbook(real)["Sheet"]["B1"].value is None


class TestExceptionsPropagateAndStillCleanUp:
    def test_a_genuine_authoring_mistake_raises_and_leaves_the_real_file_untouched(
        self, tmp_path: Path
    ) -> None:
        """write_row's positional mode without start_column raises ActionExecutionError —
        a genuine mistake, not a normal search-miss, so it must propagate, not be swallowed
        into an error StepResult."""
        real = _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: bad_write_row
                action: write_row
                workbook: manip
                sheet: "Sheet"
                row: 1
                values: ["a", "b"]
            """,
        )

        with pytest.raises(ActionExecutionError):
            run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert openpyxl.load_workbook(real)["Sheet"]["A1"].value == "hello"  # untouched


class TestValidationRunsBeforeAnyWorkbookIsTouched:
    def test_invalid_workflow_raises_before_touching_the_real_file(self, tmp_path: Path) -> None:
        real = _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: bad_step
                action: read_range
                workbook: manip
                sheet: "Sheet"
            """,
        )

        with pytest.raises(ValidationError):
            run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert openpyxl.load_workbook(real)["Sheet"]["A1"].value == "hello"


class TestCopyAcrossTwoWorkbooks:
    def test_copies_a_range_from_one_workbook_into_another(self, tmp_path: Path) -> None:
        _make_workbook(tmp_path / "output" / "historical.xlsx")
        _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              historical:
                file: "{{ env.output_folder }}/historical.xlsx"
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: copy_it
                action: copy
                source:
                  workbook: historical
                  sheet: "Sheet"
                  range: "A1"
                target:
                  workbook: manip
                  sheet: "Sheet"
                  range: "D1"
            """,
        )

        result = run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        assert result.status == "success"
        reopened = openpyxl.load_workbook(tmp_path / "output" / "manip.xlsx")
        assert reopened["Sheet"]["D1"].value == "hello"


class TestAuditLog:
    def test_audit_log_has_one_record_per_step(self, tmp_path: Path) -> None:
        _make_workbook(tmp_path / "output" / "manip.xlsx")
        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: s1
                action: read_range
                workbook: manip
                sheet: "Sheet"
                range: "A1"
              - id: s2
                action: write_cell
                workbook: manip
                sheet: "Sheet"
                cell: "B1"
                value: "x"
            """,
        )

        result = run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        lines = result.audit_log_path.read_text().splitlines()
        assert len(lines) == 2
        assert [json.loads(line)["step_id"] for line in lines] == ["s1", "s2"]


class TestCrashSafety:
    """PRD sec 6.3/6.3.1's actual crash-safety requirement: a run interrupted mid-step must
    never leave the real files touched, must leave the scratch copies in place as the
    recovery/debugging artifact, and must not leave anything in a state that blocks a later,
    valid run against the same workbook. "No orphaned Excel process" (PRD sec 6.3) isn't
    testable yet — there's no COM backend built, so no Excel process is ever spawned by the
    current (file-backend only) action set; that part of the requirement gets a real test once
    build order item 9 exists.

    A raised exception (not an ActionResult(status="error")) is what "crashes" a run, per the
    error-handling policy established in Spec sec 4/runner.py's design — write_row's positional
    mode without start_column is used here as a realistic, already-covered way to trigger one.
    """

    def test_crash_mid_run_leaves_real_file_untouched_and_scratch_copy_in_place(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        real = _make_workbook(tmp_path / "output" / "manip.xlsx")
        run_dir = tmp_path / "run_dir"
        run_dir.mkdir()
        # Deterministic run directory instead of a fresh tempfile.mkdtemp() location per run,
        # so the test can find the surviving scratch copy without globbing the system temp dir
        # (which could pick up unrelated leftovers from other test runs).
        monkeypatch.setattr("excel_runner.runner.tempfile.mkdtemp", lambda prefix=None: str(run_dir))

        workflow_path = _write_yaml(
            tmp_path / "workflow.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: write_first
                action: write_cell
                workbook: manip
                sheet: "Sheet"
                cell: "B1"
                value: "in progress"
              - id: crash_here
                action: write_row
                workbook: manip
                sheet: "Sheet"
                row: 1
                values: ["a", "b"]
            """,
        )

        with pytest.raises(ActionExecutionError):
            run_workflow(workflow_path, env_overrides={"output_folder": str(tmp_path / "output")})

        # the real file is completely untouched by the in-progress write
        assert openpyxl.load_workbook(real)["Sheet"]["B1"].value is None

        # the scratch copy survives as the recovery artifact, with the in-progress work intact
        scratch_file = run_dir / "scratch" / "manip.xlsx"
        assert scratch_file.exists()
        assert openpyxl.load_workbook(scratch_file)["Sheet"]["B1"].value == "in progress"

        # the audit log survives too — it lives outside scratch/ specifically so cleanup (which
        # never even runs on this path) couldn't take it with it (Spec sec 6.1's bug fix)
        assert (run_dir / "audit.jsonl").exists()

    def test_a_later_valid_run_against_the_same_workbook_succeeds_after_a_crash(
        self, tmp_path: Path
    ) -> None:
        """The strongest cross-platform evidence that sessions were actually closed and no
        lingering handle survives a crash: a subsequent run against the same real file just
        works. Directly detecting an OS-level file lock would be meaningful mainly on Windows
        (PRD sec 4) and isn't reliably testable on macOS, so this is the real behavior that
        matters, tested directly instead."""
        real = _make_workbook(tmp_path / "output" / "manip.xlsx")
        crashing_workflow = _write_yaml(
            tmp_path / "crash.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: crash_here
                action: write_row
                workbook: manip
                sheet: "Sheet"
                row: 1
                values: ["a", "b"]
            """,
        )
        with pytest.raises(ActionExecutionError):
            run_workflow(crashing_workflow, env_overrides={"output_folder": str(tmp_path / "output")})

        valid_workflow = _write_yaml(
            tmp_path / "valid.yaml",
            """
            env:
              output_folder: "./output"
            workbooks:
              manip:
                file: "{{ env.output_folder }}/manip.xlsx"
            steps:
              - id: write_it
                action: write_cell
                workbook: manip
                sheet: "Sheet"
                cell: "C1"
                value: "worked"
            """,
        )

        result = run_workflow(valid_workflow, env_overrides={"output_folder": str(tmp_path / "output")})

        assert result.status == "success"
        assert openpyxl.load_workbook(real)["Sheet"]["C1"].value == "worked"
