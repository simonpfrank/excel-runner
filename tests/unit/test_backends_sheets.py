"""Unit tests for backends.py's sheet-management primitives: create_sheet, rename_sheet,
delete_sheet (PRD sec 7 gap — no prior way to add/rename/remove a worksheet).
"""

from pathlib import Path

import openpyxl
import pytest

from excel_runner import backends


def _make_workbook(tmp_path: Path, name: str = "fixture.xlsx") -> Path:
    path = tmp_path / name
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "Region"
    workbook.save(path)
    return path


class TestCreateSheet:
    def test_adds_a_new_empty_sheet(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(str(_make_workbook(tmp_path)), mode="read_write")
        backends.create_sheet(workbook, "Data")
        assert "Data" in workbook.sheetnames
        workbook.close()

    def test_inserts_at_a_given_index(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(str(_make_workbook(tmp_path)), mode="read_write")
        backends.create_sheet(workbook, "Data", index=0)
        assert workbook.sheetnames[0] == "Data"
        workbook.close()

    def test_duplicate_name_raises(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(str(_make_workbook(tmp_path)), mode="read_write")
        with pytest.raises(ValueError, match="Summary"):
            backends.create_sheet(workbook, "Summary")
        workbook.close()


class TestRenameSheet:
    def test_renames_an_existing_sheet(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(str(_make_workbook(tmp_path)), mode="read_write")
        backends.rename_sheet(workbook, "Summary", "Overview")
        assert "Overview" in workbook.sheetnames
        assert "Summary" not in workbook.sheetnames
        assert workbook["Overview"]["A1"].value == "Region"
        workbook.close()


class TestDeleteSheet:
    def test_removes_a_sheet(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(str(_make_workbook(tmp_path)), mode="read_write")
        backends.create_sheet(workbook, "Data")
        backends.delete_sheet(workbook, "Data")
        assert "Data" not in workbook.sheetnames
        workbook.close()

    def test_deleting_the_only_sheet_raises(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(str(_make_workbook(tmp_path)), mode="read_write")
        with pytest.raises(ValueError, match="only sheet"):
            backends.delete_sheet(workbook, "Summary")
        workbook.close()
