"""Unit tests for the file-backend (openpyxl) primitives in excel_runner.backends (Spec sec 3).

These exercise real openpyxl against throwaway files rather than mocking it — openpyxl needs
no live Excel, so this is a real dependency being tested cheaply, same spirit as the project's
"no mocks in integration tests" convention applied at the unit level (Spec sec 7).
"""

from pathlib import Path

import openpyxl
import pytest

from excel_runner import backends
from excel_runner.backends import OwnedInstanceRegistry
from tests.unit.conftest import requires_excel, requires_working_xlwings_save


def _make_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "Region"
    sheet["B1"] = "Total"
    sheet["A2"] = "North"
    sheet["B2"] = 100
    workbook.save(path)
    return path


class TestOpenWorkbook:
    def test_opens_read_write_by_default(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        workbook = backends.open_workbook(str(path), mode="read_write")
        assert workbook["Summary"]["A1"].value == "Region"
        workbook.close()

    def test_opens_read_only(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        workbook = backends.open_workbook(str(path), mode="read_only")
        assert workbook["Summary"]["A1"].value == "Region"
        workbook.close()


def _make_formula_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "formula_fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = 10
    sheet["B1"] = "=A1*2"
    workbook.save(path)
    return path


class TestOpenWorkbookDataOnlyDefault:
    """`data_only` — not `read_only` — is the sole determinant of formula-vs-value reads
    (confirmed via runtime probing). A freshly-written, never-recalculated formula has no
    cached value at all, so this doesn't need a real Excel instance to prove the default
    actually flipped to data_only=True: reading returns None (cache absent), not the formula
    text, which is exactly what data_only=False would have returned instead."""

    def test_defaults_to_data_only_true(self, tmp_path: Path) -> None:
        path = _make_formula_workbook(tmp_path)
        workbook = backends.open_workbook(str(path), mode="read_only")
        assert backends.read_range(workbook, "Summary", "B1") is None
        workbook.close()

    def test_data_only_false_returns_the_formula_text(self, tmp_path: Path) -> None:
        path = _make_formula_workbook(tmp_path)
        workbook = backends.open_workbook(str(path), mode="read_only", data_only=False)
        assert backends.read_range(workbook, "Summary", "B1") == "=A1*2"
        workbook.close()


@requires_excel
@requires_working_xlwings_save
class TestOpenWorkbookForFormulaRead:
    def test_reads_the_formula_text_even_when_a_real_cached_value_exists(
        self, tmp_path: Path
    ) -> None:
        path = _make_formula_workbook(tmp_path)
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            backends.com_calculate_workbook(book)
            backends.com_wait_until_calculation_done(app)
            backends.xlw_save_workbook(book)
            backends.xlw_close_workbook(book)
        finally:
            registry.close_owned()

        # The default open now returns the real cached value...
        default_workbook = backends.open_workbook(str(path), mode="read_only")
        assert backends.read_range(default_workbook, "Summary", "B1") == 20
        default_workbook.close()

        # ...while the formula-view helper returns the formula text instead.
        formula_workbook = backends.open_workbook_for_formula_read(str(path))
        assert backends.read_range(formula_workbook, "Summary", "B1") == "=A1*2"
        formula_workbook.close()


class TestReadRange:
    def test_reads_a_single_cell_as_a_scalar(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        assert backends.read_range(workbook, "Summary", "A1") == "Region"
        workbook.close()

    def test_reads_a_multi_cell_range_as_a_2d_list(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        assert backends.read_range(workbook, "Summary", "A1:B2") == [
            ["Region", "Total"],
            ["North", 100],
        ]
        workbook.close()

    def test_reads_a_single_row_range_as_a_2d_list(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        assert backends.read_range(workbook, "Summary", "A1:B1") == [
            ["Region", "Total"]
        ]
        workbook.close()


def _make_named_range_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "named_range_fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "Region"
    sheet["B1"] = "Total"
    sheet["A2"] = "North"
    sheet["B2"] = 100
    workbook.defined_names["SalesTotal"] = openpyxl.workbook.defined_name.DefinedName(
        "SalesTotal", attr_text="Summary!$B$2"
    )
    workbook.save(path)
    return path


class TestResolveRange:
    def test_plain_a1_notation_passes_through_unchanged(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        assert backends.resolve_range(workbook, "Summary", "A1:B2") == (
            "Summary",
            "A1:B2",
        )
        workbook.close()

    def test_resolves_a_defined_name_to_its_own_sheet_and_a1_range(
        self, tmp_path: Path
    ) -> None:
        workbook = backends.open_workbook(
            str(_make_named_range_workbook(tmp_path)), mode="read_write"
        )
        assert backends.resolve_range(workbook, "Summary", "SalesTotal") == (
            "Summary",
            "B2",
        )
        workbook.close()

    def test_a_defined_names_own_sheet_wins_over_the_passed_sheet(
        self, tmp_path: Path
    ) -> None:
        workbook = backends.open_workbook(
            str(_make_named_range_workbook(tmp_path)), mode="read_write"
        )
        assert backends.resolve_range(workbook, "SomeOtherSheet", "SalesTotal") == (
            "Summary",
            "B2",
        )
        workbook.close()


class TestReadRangeNamedRange:
    def test_reads_the_value_at_a_defined_name(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_named_range_workbook(tmp_path)), mode="read_write"
        )
        assert backends.read_range(workbook, "Summary", "SalesTotal") == 100
        workbook.close()


class TestResolveSheetNames:
    """PRD sec 7's read_range row: sheet can be a single name, an explicit list, "all"
    (every sheet), or {"matching": <regex>} (every sheet whose name matches, re.search-style
    like find_row/find_headers_row's `patterns`) -- all resolved to an explicit list here,
    one code path underneath either way."""

    def test_single_string_returns_a_one_item_list(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        assert backends.resolve_sheet_names(workbook, "Summary") == ["Summary"]
        workbook.close()

    def test_a_list_is_returned_as_is(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        assert backends.resolve_sheet_names(workbook, ["Summary"]) == ["Summary"]
        workbook.close()

    def test_all_expands_to_every_sheet_in_workbook_order(self, tmp_path: Path) -> None:
        path = tmp_path / "multi.xlsx"
        workbook = openpyxl.Workbook()
        first = workbook.active
        assert first is not None
        first.title = "First"
        workbook.create_sheet("Second")
        workbook.save(path)
        workbook.close()

        reopened = backends.open_workbook(str(path), mode="read_write")
        assert backends.resolve_sheet_names(reopened, "all") == ["First", "Second"]
        reopened.close()

    def test_matching_selects_sheets_whose_name_matches_the_regex(
        self, tmp_path: Path
    ) -> None:
        path = tmp_path / "multi.xlsx"
        workbook = openpyxl.Workbook()
        first = workbook.active
        assert first is not None
        first.title = "A&H North"
        workbook.create_sheet("A&H South")
        workbook.create_sheet("Other")
        workbook.save(path)
        workbook.close()

        reopened = backends.open_workbook(str(path), mode="read_write")
        assert backends.resolve_sheet_names(reopened, {"matching": "^A&H"}) == [
            "A&H North",
            "A&H South",
        ]
        reopened.close()


class TestWriteCell:
    def test_writes_a_value(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        backends.write_cell(workbook, "Summary", "C1", "Status")
        assert workbook["Summary"]["C1"].value == "Status"
        workbook.close()

    def test_writes_a_formula_string_as_is(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        backends.write_cell(workbook, "Summary", "C2", "=SUM(B2:B2)")
        assert workbook["Summary"]["C2"].value == "=SUM(B2:B2)"
        workbook.close()


class TestSaveWorkbook:
    def test_saves_changes_to_the_given_path(self, tmp_path: Path) -> None:
        source = _make_workbook(tmp_path)
        workbook = backends.open_workbook(str(source), mode="read_write")
        backends.write_cell(workbook, "Summary", "A1", "Changed")
        out_path = tmp_path / "saved.xlsx"
        backends.save_workbook(workbook, str(out_path))
        workbook.close()

        reopened = backends.open_workbook(str(out_path), mode="read_only")
        assert backends.read_range(reopened, "Summary", "A1") == "Changed"
        reopened.close()


class TestCloseWorkbook:
    def test_close_does_not_raise(self, tmp_path: Path) -> None:
        workbook = backends.open_workbook(
            str(_make_workbook(tmp_path)), mode="read_write"
        )
        backends.close_workbook(workbook)  # should not raise

    def test_open_missing_file_raises_a_clear_file_not_found_error(
        self, tmp_path: Path
    ) -> None:
        with pytest.raises(FileNotFoundError):
            backends.open_workbook(
                str(tmp_path / "does_not_exist.xlsx"), mode="read_write"
            )
