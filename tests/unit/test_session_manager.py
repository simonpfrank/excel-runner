"""Unit tests for SessionManager (Spec sec 5.2).

Read/write mode is caller-specified here, not statically inferred — tier-2 validation
(Spec sec 5.4, not built yet) will compute that later and hand it in; SessionManager itself
doesn't guess. A session opened with mode="read_write" is treated as "this run will write to
it" and gets staged through ScratchManager (PRD sec 6.3.1); mode="read_only" is now staged too
(PRD sec 6.2.3's correction — avoids holding a handle open on the real file), just never
committed back since nothing about it ever changes.
"""

import logging
import shutil
from pathlib import Path

import openpyxl
import pytest

from excel_runner import backends, engine
from excel_runner.core import ActionExecutionError, WorkbookRef, WorkbookSession
from excel_runner.engine import ScratchManager, SessionManager
from tests.unit.conftest import requires_excel, requires_working_xlwings_save


def _write_workbook(path: Path, cell_value: str = "original") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["A1"] = cell_value
    workbook.save(path)
    return path


def _make_target_and_linking_workbooks(
    tmp_path: Path, target_value: float = 5
) -> tuple[Path, Path]:
    """A real xlsx pair with a genuine R4 (absolute-path) external link between them, in two
    separate real folders. Built the same way the plan doc's probes did: `linking` first gets
    a same-folder relative link (so the formula itself resolves), is `ChangeLink`'d to the
    target's absolute path, then gets moved to its own separate real folder — reproducing
    exactly what an R4 link looks like on disk, without needing to hand-author rels XML.

    Returns:
        (target_path, linking_path) — both real, on-disk, closed workbooks.
    """
    registry = backends.OwnedInstanceRegistry()
    app = registry.spawn()
    try:
        target = app.books.add()
        target.sheets[0].range("A1").value = target_value
        target_path = tmp_path / "target_dir" / "target.xlsx"
        target_path.parent.mkdir(parents=True, exist_ok=True)
        target.save(str(target_path))

        linking = app.books.add()
        linking.sheets[0].range("A1").formula = "='[target.xlsx]Sheet1'!A1*2"
        same_folder_path = tmp_path / "target_dir" / "linking.xlsx"
        linking.save(str(same_folder_path))
        linking.api.ChangeLink(
            Name="target.xlsx", NewName=str(target_path.resolve()), Type=1
        )
        linking.save()
        linking.close()
        target.close()
    finally:
        registry.close_owned()

    linking_path = tmp_path / "linking_dir" / "linking.xlsx"
    linking_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(same_folder_path), str(linking_path))
    return target_path, linking_path


class TestGetOrOpen:
    def test_opens_an_existing_workbook_read_write_by_default(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("manip")

        assert session.name == "manip"
        assert session.mode == "read_write"
        assert session.handle["Sheet"]["A1"].value == "original"

    def test_read_write_stages_through_scratch(self, tmp_path: Path) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("manip", mode="read_write")

        assert session.scratch_path is not None
        assert str(tmp_path / "working" / "scratch") in session.path
        assert session.path != str(real)

    def test_read_only_stages_too_but_is_never_committed(self, tmp_path: Path) -> None:
        """PRD sec 6.2.3's correction: read-only sessions are staged like read-write ones now
        (avoids holding a handle open on the real file), just never committed back."""
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("manip", mode="read_only")

        assert session.scratch_path is not None
        assert str(tmp_path / "working" / "scratch") in session.path
        assert session.path != str(real)

    def test_second_call_for_the_same_name_returns_the_cached_session(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        first = manager.get_or_open("manip")
        second = manager.get_or_open("manip")

        assert first is second

    def test_unknown_workbook_name_raises_a_clear_error(self, tmp_path: Path) -> None:
        manager = SessionManager({}, ScratchManager(tmp_path / "working"))
        with pytest.raises(ActionExecutionError) as exc_info:
            manager.get_or_open("nonexistent")
        assert "nonexistent" in exc_info.value.detail.message

    def test_missing_file_without_create_if_missing_raises_a_clear_error(
        self, tmp_path: Path
    ) -> None:
        workbooks = {
            "manip": WorkbookRef(name="manip", file=str(tmp_path / "missing.xlsx"))
        }
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        with pytest.raises(ActionExecutionError) as exc_info:
            manager.get_or_open("manip")
        assert "missing.xlsx" in exc_info.value.detail.message


class TestNeededBackend:
    """PRD sec 6.2.2's capability -> backend mapping, as a standalone pure function."""

    def test_file_capability_needs_file_backend(self) -> None:
        assert engine._needed_backend("file") == "file"

    def test_xlw_capability_needs_xlw_backend(self) -> None:
        assert engine._needed_backend("xlw") == "xlw"

    def test_com_capability_needs_xlw_backend_too(self) -> None:
        """com reaches deeper via xlwings' .api on an xlw-backed session — SessionManager
        never needs a distinct backend state for it (PRD sec 6.2.2)."""
        assert engine._needed_backend("com") == "xlw"

    def test_depends_on_param_capability_is_not_resolvable_here(self) -> None:
        """read_metadata's real capability depends on its target: param at runtime — that
        resolution isn't built yet (Spec sec 5.1), so this can't be mapped to a backend yet.
        """
        with pytest.raises(ActionExecutionError):
            engine._needed_backend("depends_on_param")

    def test_none_capability_is_not_resolvable_here(self) -> None:
        """Control actions (stop) never reach get_or_open at all — no workbook: field — so this
        is defensive, not a real path (PRD sec 6.9)."""
        with pytest.raises(ActionExecutionError):
            engine._needed_backend("none")


class TestCapabilityBackendMatch:
    def test_matching_capability_returns_the_session_normally(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("manip", capability="file")

        assert session.backend == "file"

    def test_default_capability_is_file_unchanged_from_before_this_param_existed(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("manip")  # no capability given

        assert session.backend == "file"


@requires_excel
class TestBackendSwitching:
    """PRD sec 6.2.2: bidirectional backend switching, against a real Excel instance — no
    mocks (project convention, matches test_owned_instance_registry.py)."""

    def test_brand_new_session_opens_directly_on_the_needed_backend(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        try:
            session = manager.get_or_open("manip", capability="xlw")
            assert session.backend == "xlw"
            assert session.handle.name == "manip.xlsx"
        finally:
            manager.close_all()

    def test_switching_backend_logs_at_info(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        manager.get_or_open("manip", capability="file")

        try:
            with caplog.at_level(logging.INFO, logger="excel_runner.engine"):
                manager.get_or_open("manip", capability="xlw")
        finally:
            manager.close_all()

        messages = " ".join(r.message for r in caplog.records)
        assert "manip" in messages
        assert "file" in messages
        assert "xlw" in messages

    def test_switching_an_open_file_session_to_xlw_reopens_it_there(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        file_session = manager.get_or_open("manip", capability="file")
        file_session.handle["Sheet"]["A1"] = "written on file backend"
        file_session.dirty = True

        try:
            xlw_session = manager.get_or_open("manip", capability="xlw")
            assert xlw_session is file_session  # same session object, mutated in place
            assert xlw_session.backend == "xlw"
            assert (
                xlw_session.handle.sheets["Sheet"]["A1"].value
                == "written on file backend"
            )
        finally:
            manager.close_all()

    def test_switching_an_open_xlw_session_back_to_file_reopens_it_there(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        xlw_session = manager.get_or_open("manip", capability="xlw")
        xlw_session.handle.sheets["Sheet"]["A1"].value = "written on xlw backend"
        xlw_session.dirty = True

        try:
            file_session = manager.get_or_open("manip", capability="file")
            assert file_session is xlw_session
            assert file_session.backend == "file"
            assert file_session.handle["Sheet"]["A1"].value == "written on xlw backend"
        finally:
            manager.close_all()

    def test_two_workbooks_needing_xlw_share_one_excel_instance(
        self, tmp_path: Path
    ) -> None:
        real_a = _write_workbook(tmp_path / "real" / "a.xlsx")
        real_b = _write_workbook(tmp_path / "real" / "b.xlsx")
        workbooks = {
            "a": WorkbookRef(name="a", file=str(real_a)),
            "b": WorkbookRef(name="b", file=str(real_b)),
        }
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        try:
            session_a = manager.get_or_open("a", capability="xlw")
            session_b = manager.get_or_open("b", capability="xlw")
            assert session_a.handle.app.pid == session_b.handle.app.pid
        finally:
            manager.close_all()

    def test_close_all_quits_the_shared_owned_excel_instance(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        manager.get_or_open("manip", capability="xlw")

        manager.close_all()

        assert manager._owned_instances.pids == ()

    def test_close_all_is_a_no_op_for_the_owned_instance_when_xlw_was_never_needed(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        manager.get_or_open("manip", capability="file")

        manager.close_all()  # should not raise, nothing xlw-related was ever spawned

        assert manager._owned_instances.pids == ()


class TestCreateIfMissing:
    def test_read_only_with_create_if_missing_creates_at_the_scratch_path(
        self, tmp_path: Path
    ) -> None:
        """Unusual combination (why read a workbook you just created blank?) but not
        forbidden — must still work correctly rather than being an untested code path.
        Creates at the scratch path now, not the real path (PRD sec 6.2.3's correction —
        read-only is staged too)."""
        real = tmp_path / "real" / "new.xlsx"
        workbooks = {
            "new": WorkbookRef(name="new", file=str(real), create_if_missing=True)
        }
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("new", mode="read_only")

        assert not real.exists()
        assert session.scratch_path is not None
        assert session.handle.sheetnames

    def test_creates_a_blank_workbook_at_the_scratch_path(self, tmp_path: Path) -> None:
        real = tmp_path / "real" / "new.xlsx"
        workbooks = {
            "new": WorkbookRef(name="new", file=str(real), create_if_missing=True)
        }
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("new")

        assert not real.exists()  # not created in place — only in scratch, until commit
        assert session.handle.sheetnames

    def test_creates_from_a_template_workbook(self, tmp_path: Path) -> None:
        template_real = _write_workbook(
            tmp_path / "real" / "historical.xlsx", "template content"
        )
        new_real = tmp_path / "real" / "results.xlsx"
        workbooks = {
            "historical": WorkbookRef(name="historical", file=str(template_real)),
            "results": WorkbookRef(
                name="results",
                file=str(new_real),
                create_if_missing=True,
                template="historical",
            ),
        }
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))

        session = manager.get_or_open("results")

        assert session.handle["Sheet"]["A1"].value == "template content"


class TestCloseAll:
    def test_closes_every_opened_session(self, tmp_path: Path) -> None:
        real_a = _write_workbook(tmp_path / "real" / "a.xlsx")
        real_b = _write_workbook(tmp_path / "real" / "b.xlsx")
        workbooks = {
            "a": WorkbookRef(name="a", file=str(real_a)),
            "b": WorkbookRef(name="b", file=str(real_b)),
        }
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        manager.get_or_open("a")
        manager.get_or_open("b")

        manager.close_all()  # should not raise

    def test_close_all_with_no_sessions_opened_does_not_raise(
        self, tmp_path: Path
    ) -> None:
        manager = SessionManager({}, ScratchManager(tmp_path / "working"))
        manager.close_all()

    def test_one_failing_close_does_not_prevent_others_from_closing(
        self, tmp_path: Path
    ) -> None:
        """Crash-safety requirement (PRD sec 6.3): every session must get a close attempt,
        even if an earlier one fails. Uses a fake handle whose close() raises — openpyxl's
        own Workbook.close() is a no-op even when called twice, so it can't produce a real
        failure to test against."""

        class _ExplodingHandle:
            def close(self) -> None:
                raise RuntimeError("simulated close failure")

        real_b = _write_workbook(tmp_path / "real" / "b.xlsx")
        workbooks = {"b": WorkbookRef(name="b", file=str(real_b))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        # "a" (which will fail to close) is inserted first, so iteration reaches it before "b" —
        # this is what actually proves close_all() doesn't stop after the first failure.
        manager._sessions["a"] = WorkbookSession(
            name="a",
            backend="file",
            handle=_ExplodingHandle(),
            path="a.xlsx",
            mode="read_write",
        )
        session_b = manager.get_or_open("b")
        closed_b = False
        real_close_b = session_b.handle.close

        def _tracking_close() -> None:
            nonlocal closed_b
            closed_b = True
            real_close_b()

        session_b.handle.close = _tracking_close

        with pytest.raises(ExceptionGroup):
            manager.close_all()

        assert closed_b is True


class TestCheckpoint:
    """checkpoint() persists in-progress writes to the scratch file mid-run, so a later crash
    leaves everything that succeeded so far visible in the recovery artifact (PRD sec 6.3.1) —
    found necessary via a failing integration test: without this, the scratch file on disk
    only ever reflected whatever was there at staging time, since openpyxl writes stay in
    memory until an explicit save and nothing else triggers one mid-run."""

    def test_checkpoint_saves_a_dirty_staged_session_to_its_scratch_file(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        session = manager.get_or_open("manip", mode="read_write")
        session.handle["Sheet"]["A1"] = "in progress"
        session.dirty = True

        manager.checkpoint()

        assert (
            openpyxl.load_workbook(session.path)["Sheet"]["A1"].value == "in progress"
        )

    def test_checkpoint_does_not_touch_the_real_path(self, tmp_path: Path) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        session = manager.get_or_open("manip", mode="read_write")
        session.handle["Sheet"]["A1"] = "in progress"
        session.dirty = True

        manager.checkpoint()

        assert openpyxl.load_workbook(real)["Sheet"]["A1"].value == "original"

    def test_checkpoint_clears_the_dirty_flag(self, tmp_path: Path) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        session = manager.get_or_open("manip", mode="read_write")
        session.dirty = True

        manager.checkpoint()

        assert session.dirty is False

    def test_checkpoint_skips_a_non_dirty_session(self, tmp_path: Path) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        manager.get_or_open("manip", mode="read_write")

        manager.checkpoint()  # should not raise

    def test_checkpoint_skips_read_only_sessions(self, tmp_path: Path) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        manager.get_or_open("manip", mode="read_only")

        manager.checkpoint()  # should not raise


class TestCommitAll:
    def test_saves_dirty_staged_sessions_and_commits_them_to_the_real_path(
        self, tmp_path: Path
    ) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        session = manager.get_or_open("manip", mode="read_write")
        session.handle["Sheet"]["A1"] = "changed"
        session.dirty = True

        manager.commit_all()

        reopened = openpyxl.load_workbook(real)
        assert reopened["Sheet"]["A1"].value == "changed"

    def test_read_only_sessions_are_not_touched_by_commit(self, tmp_path: Path) -> None:
        real = _write_workbook(tmp_path / "real" / "manip.xlsx")
        workbooks = {"manip": WorkbookRef(name="manip", file=str(real))}
        manager = SessionManager(workbooks, ScratchManager(tmp_path / "working"))
        manager.get_or_open("manip", mode="read_only")

        manager.commit_all()  # should not raise, nothing to commit


@requires_excel
@requires_working_xlwings_save
class TestR4LinkWiringAtStaging:
    """docs/recalc_and_link_refresh_plan.md sec 2 R4.1: staging-time ChangeLink, real Excel
    throughout (project convention, matches TestBackendSwitching/TestLinkPrimitives)."""

    def test_staging_both_sides_repoints_the_link_to_the_targets_scratch_copy(
        self, tmp_path: Path
    ) -> None:
        target_path, linking_path = _make_target_and_linking_workbooks(
            tmp_path, target_value=5
        )
        workbooks = {
            "target": WorkbookRef(name="target", file=str(target_path)),
            "linking": WorkbookRef(name="linking", file=str(linking_path)),
        }
        link_targets = {"linking": {"target"}, "target": set()}
        manager = SessionManager(
            workbooks, ScratchManager(tmp_path / "working"), link_targets=link_targets
        )
        try:
            manager.get_or_open("target", capability="xlw")
            # Diverge the real target file now that its scratch copy already holds 5 — a
            # value only visible on the linking side if its link is NOT still pointing here.
            stale = openpyxl.load_workbook(target_path)
            stale.active["A1"] = 999
            stale.save(target_path)

            linking_session = manager.get_or_open("linking", capability="xlw")

            assert ("linking", "target") in manager._wired_r4_links
            assert (
                linking_session.handle.sheets[0].range("A1").value == 10
            )  # 5 * 2, from scratch
        finally:
            manager.close_all()

    def test_wiring_a_pair_only_happens_once(self, tmp_path: Path) -> None:
        target_path, linking_path = _make_target_and_linking_workbooks(tmp_path)
        workbooks = {
            "target": WorkbookRef(name="target", file=str(target_path)),
            "linking": WorkbookRef(name="linking", file=str(linking_path)),
        }
        link_targets = {"linking": {"target"}, "target": set()}
        manager = SessionManager(
            workbooks, ScratchManager(tmp_path / "working"), link_targets=link_targets
        )
        try:
            manager.get_or_open("target", capability="xlw")
            manager.get_or_open("linking", capability="xlw")
            manager.get_or_open("linking", capability="xlw")  # cached, second call
            manager.get_or_open("target", capability="xlw")  # cached, second call

            assert manager._wired_r4_links == {("linking", "target")}
        finally:
            manager.close_all()

    def test_wiring_logs_at_info(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        target_path, linking_path = _make_target_and_linking_workbooks(
            tmp_path, target_value=5
        )
        workbooks = {
            "target": WorkbookRef(name="target", file=str(target_path)),
            "linking": WorkbookRef(name="linking", file=str(linking_path)),
        }
        link_targets = {"linking": {"target"}, "target": set()}
        manager = SessionManager(
            workbooks, ScratchManager(tmp_path / "working"), link_targets=link_targets
        )
        try:
            manager.get_or_open("target", capability="xlw")
            with caplog.at_level(logging.INFO, logger="excel_runner.engine"):
                manager.get_or_open("linking", capability="xlw")
        finally:
            manager.close_all()

        messages = " ".join(r.message for r in caplog.records)
        assert "linking" in messages
        assert "target" in messages


@requires_excel
@requires_working_xlwings_save
class TestR4LinkRevertAndCommit:
    """docs/recalc_and_link_refresh_plan.md sec 3.2.1: commit-time ChangeLink revert (target
    scratch path -> target real path), then save — once the target is already committed
    (R5 commit order), before `name`'s own scratch-to-real commit."""

    def test_commit_all_reverts_the_link_and_both_real_files_end_up_correct(
        self, tmp_path: Path
    ) -> None:
        target_path, linking_path = _make_target_and_linking_workbooks(
            tmp_path, target_value=5
        )
        workbooks = {
            "target": WorkbookRef(name="target", file=str(target_path)),
            "linking": WorkbookRef(name="linking", file=str(linking_path)),
        }
        link_targets = {"linking": {"target"}, "target": set()}
        commit_order = engine.compute_link_commit_order(link_targets)
        manager = SessionManager(
            workbooks,
            ScratchManager(tmp_path / "working"),
            link_targets=link_targets,
            commit_order=commit_order,
        )
        try:
            target_session = manager.get_or_open("target", capability="xlw")
            manager.get_or_open(
                "linking", capability="xlw"
            )  # triggers staging-time wiring

            target_session.handle.sheets[0].range("A1").value = 50
            target_session.dirty = True
            manager.checkpoint()

            manager.commit_all()
        finally:
            manager.close_all()

        # target's real file now holds its final, committed content
        assert openpyxl.load_workbook(target_path)["Sheet1"]["A1"].value == 50

        # linking's real file's stored link is back to target's real, absolute path — not left
        # pointing at a scratch copy that won't exist once this run's scratch dir is gone
        raw_targets = engine.scan_external_link_targets(linking_path)
        resolved = {engine.resolve_link_target(t, linking_path) for t in raw_targets}
        assert target_path.resolve() in resolved

        # and its cached formula value reflects target's final, real content
        cached = openpyxl.load_workbook(linking_path, data_only=True)["Sheet1"][
            "A1"
        ].value
        assert cached == 100  # 50 * 2

    def test_revert_logs_at_info(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        target_path, linking_path = _make_target_and_linking_workbooks(
            tmp_path, target_value=5
        )
        workbooks = {
            "target": WorkbookRef(name="target", file=str(target_path)),
            "linking": WorkbookRef(name="linking", file=str(linking_path)),
        }
        link_targets = {"linking": {"target"}, "target": set()}
        commit_order = engine.compute_link_commit_order(link_targets)
        manager = SessionManager(
            workbooks,
            ScratchManager(tmp_path / "working"),
            link_targets=link_targets,
            commit_order=commit_order,
        )
        try:
            manager.get_or_open("target", capability="xlw")
            manager.get_or_open("linking", capability="xlw")

            with caplog.at_level(logging.INFO, logger="excel_runner.engine"):
                manager.commit_all()
        finally:
            manager.close_all()

        messages = " ".join(r.message for r in caplog.records)
        assert "linking" in messages
        assert "target" in messages
