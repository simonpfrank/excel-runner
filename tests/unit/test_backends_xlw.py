"""Unit tests for the xlwings (live-Excel) backend primitives in excel_runner.backends
(Spec sec 3). Real xlwings against a real, locally-spawned Excel instance — no mocks (project
convention) — same spirit as test_owned_instance_registry.py.
"""

from pathlib import Path

import openpyxl
import pytest
import xlwings as xw

from excel_runner import backends
from excel_runner.backends import OwnedInstanceRegistry
from tests.unit.conftest import requires_excel, requires_working_xlwings_save


def _make_workbook(path: Path) -> Path:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "hello"
    workbook.save(path)
    return path


@requires_excel
class TestXlwOpenWorkbook:
    def test_opens_an_existing_workbook(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            assert book.name == "fixture.xlsx"
            assert book.sheets["Summary"]["A1"].value == "hello"
        finally:
            registry.close_owned()

    def test_missing_file_raises_file_not_found(self, tmp_path: Path) -> None:
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            with pytest.raises(FileNotFoundError):
                backends.xlw_open_workbook(
                    app, str(tmp_path / "does_not_exist.xlsx"), mode="read_write"
                )
        finally:
            registry.close_owned()


@requires_excel
class TestXlwCloseWorkbook:
    def test_closes_the_book_without_quitting_the_app(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            backends.xlw_close_workbook(book)
            # spawn() spawns with add_book=True (a Windows reliability fix, see
            # OwnedInstanceRegistry.spawn's docstring) — the app's own initial "Book1"
            # is expected to remain; the one this test opened and closed must not.
            assert "fixture.xlsx" not in [b.name for b in app.books]
        finally:
            registry.close_owned()


@requires_excel
@requires_working_xlwings_save
class TestXlwSaveWorkbook:
    def test_saves_in_place_and_the_write_persists(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            book.sheets["Summary"]["B1"].value = "written via xlwings"
            backends.xlw_save_workbook(book)
            backends.xlw_close_workbook(book)
        finally:
            registry.close_owned()

        reopened = openpyxl.load_workbook(path)
        assert reopened["Summary"]["B1"].value == "written via xlwings"


def _make_formula_workbook(path: Path) -> Path:
    """A workbook whose B1 depends on A1, with the cached formula result deliberately stale
    (openpyxl can't compute formulas itself, so it never gets to compute a correct one) — a
    real recalculation is what's needed to make B1 match A1*2."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = 10
    sheet["B1"] = "=A1*2"
    workbook.save(path)
    return path


@requires_excel
@requires_working_xlwings_save
class TestRecalculatePrimitives:
    def test_com_calculate_workbook_updates_a_formula(self, tmp_path: Path) -> None:
        path = _make_formula_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            book.sheets["Summary"]["A1"].value = 21
            backends.com_calculate_workbook(book)
            backends.com_wait_until_calculation_done(app)
            assert book.sheets["Summary"]["B1"].value == 42
        finally:
            registry.close_owned()

    def test_com_calculate_sheet_updates_a_formula(self, tmp_path: Path) -> None:
        path = _make_formula_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            book.sheets["Summary"]["A1"].value = 5
            backends.com_calculate_sheet(book, "Summary")
            backends.com_wait_until_calculation_done(app)
            assert book.sheets["Summary"]["B1"].value == 10
        finally:
            registry.close_owned()

    def test_xlw_calculate_all_updates_a_formula(self, tmp_path: Path) -> None:
        path = _make_formula_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            book.sheets["Summary"]["A1"].value = 3
            backends.xlw_calculate_all(app)
            backends.com_wait_until_calculation_done(app)
            assert book.sheets["Summary"]["B1"].value == 6
        finally:
            registry.close_owned()

    def test_com_calculate_full_updates_a_formula(self, tmp_path: Path) -> None:
        path = _make_formula_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            book.sheets["Summary"]["A1"].value = 7
            backends.com_calculate_full(app)
            backends.com_wait_until_calculation_done(app)
            assert book.sheets["Summary"]["B1"].value == 14
        finally:
            registry.close_owned()

    def test_com_calculate_full_rebuild_updates_a_formula(self, tmp_path: Path) -> None:
        path = _make_formula_workbook(tmp_path / "fixture.xlsx")
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = backends.xlw_open_workbook(app, str(path), mode="read_write")
            book.sheets["Summary"]["A1"].value = 9
            backends.com_calculate_full_rebuild(app)
            backends.com_wait_until_calculation_done(app)
            assert book.sheets["Summary"]["B1"].value == 18
        finally:
            registry.close_owned()

    def test_wait_raises_timeout_error_if_calculation_never_finishes(
        self, tmp_path: Path
    ) -> None:
        """A fake app whose CalculationState never reports done — proves the timeout path
        is real, without needing an actual multi-minute Excel calculation."""

        class _AlwaysCalculatingApi:
            CalculationState = 1  # xlCalculating, never changes

        class _FakeApp:
            api = _AlwaysCalculatingApi()

        with pytest.raises(TimeoutError):
            backends.com_wait_until_calculation_done(_FakeApp(), timeout=0.05)  # type: ignore[arg-type]


def _make_link_target(app: xw.App, path: Path, value: float) -> Path:
    """A standalone workbook with a single value, saved and closed — a genuine link source."""
    book = app.books.add()
    book.sheets[0].range("A1").value = value
    book.save(str(path))
    book.close()
    return path


@requires_excel
@requires_working_xlwings_save
class TestLinkPrimitives:
    """docs/recalc_and_link_refresh_plan.md sec 2 (R1/R4) — the raw COM link primitives that
    plan depends on. Real xlwings/Excel throughout, matching project convention."""

    def test_com_link_sources_lists_the_external_link(self, tmp_path: Path) -> None:
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            _make_link_target(app, tmp_path / "target.xlsx", 5)
            linking = app.books.add()
            linking.sheets[0].range("A1").formula = "='[target.xlsx]Sheet1'!A1*2"
            linking.save(str(tmp_path / "linking.xlsx"))

            assert backends.com_link_sources(linking) == ["target.xlsx"]
        finally:
            registry.close_owned()

    def test_com_link_sources_is_empty_with_no_external_links(
        self, tmp_path: Path
    ) -> None:
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            book = app.books.add()
            book.sheets[0].range("A1").value = "no links here"
            book.save(str(tmp_path / "standalone.xlsx"))

            assert backends.com_link_sources(book) == []
        finally:
            registry.close_owned()

    def test_com_change_link_repoints_and_instantly_refreshes_from_an_existing_target(
        self, tmp_path: Path
    ) -> None:
        """Matches probe9/probe7's finding: ChangeLink to a target that exists on disk
        re-evaluates dependent cells against its current content immediately, no separate
        recalculate/UpdateLink call needed."""
        registry = OwnedInstanceRegistry()
        app = registry.spawn()
        try:
            _make_link_target(app, tmp_path / "old.xlsx", 5)
            new_target = _make_link_target(app, tmp_path / "new.xlsx", 100)

            linking = app.books.add()
            linking.sheets[0].range("A1").formula = "='[old.xlsx]Sheet1'!A1*2"
            linking.save(str(tmp_path / "linking.xlsx"))

            backends.com_change_link(linking, "old.xlsx", str(new_target))

            assert linking.sheets[0].range("A1").value == 200
        finally:
            registry.close_owned()

    def test_com_update_link_reads_a_closed_workbook_from_disk(
        self, tmp_path: Path
    ) -> None:
        """Matches probe6b's finding: UpdateLink refreshes from a file this process never
        opened at all, as long as it's genuinely closed and current on disk."""
        setup_registry = OwnedInstanceRegistry()
        setup_app = setup_registry.spawn()
        try:
            target_path = _make_link_target(setup_app, tmp_path / "target.xlsx", 5)

            # Absolute reference (R4's case), not a same-folder relative one (R1) — the bare
            # filename form Excel stores for same-folder links can't always be resolved by a
            # separate app instance with a different default working folder.
            linking = setup_app.books.add()
            linking.sheets[0].range(
                "A1"
            ).formula = f"='{target_path.parent}\\[{target_path.name}]Sheet1'!A1*2"
            linking.save(str(tmp_path / "linking.xlsx"))
            linking.close()
        finally:
            setup_registry.close_owned()

        # Edit and close the target directly, in a separate, already-closed session.
        editor_registry = OwnedInstanceRegistry()
        editor_app = editor_registry.spawn()
        try:
            editor_book = backends.xlw_open_workbook(
                editor_app, str(target_path), mode="read_write"
            )
            editor_book.sheets[0].range("A1").value = 999
            backends.xlw_save_workbook(editor_book)
            backends.xlw_close_workbook(editor_book)
        finally:
            editor_registry.close_owned()

        # A third, separate app instance never opens target.xlsx directly at all.
        main_registry = OwnedInstanceRegistry()
        main_app = main_registry.spawn()
        try:
            linking = backends.xlw_open_workbook(
                main_app, str(tmp_path / "linking.xlsx"), mode="read_write"
            )
            assert (
                linking.sheets[0].range("A1").value != 1998
            )  # not yet the fresh value

            (source,) = backends.com_link_sources(linking)
            backends.com_update_link(linking, source)

            assert linking.sheets[0].range("A1").value == 1998
        finally:
            main_registry.close_owned()
