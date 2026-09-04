"""Unit tests for tier-3 existence validation (`engine.validate_existence`, opt-in via CLI
`--check-existence`). Unlike tiers 1/2, this one really opens workbooks — read-only, via
openpyxl — to confirm every sheet/defined name a step references by literal name actually
exists.
"""

from pathlib import Path

import openpyxl
import pytest

from excel_runner import engine as validation
from excel_runner.core import Step, ValidationError, WorkbookRef, Workflow


def _workflow(steps: list[Step], workbooks: dict[str, WorkbookRef]) -> Workflow:
    return Workflow(env={}, workbooks=workbooks, steps=tuple(steps))


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Products"
    sheet["A1"] = "Header"
    workbook.defined_names.add(
        openpyxl.workbook.defined_name.DefinedName("MyRange", attr_text="Products!$A$1")
    )
    workbook.save(path)
    return path


class TestSheetExistence:
    def test_passes_when_sheet_exists(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="read_range",
                    params={"workbook": "wb", "sheet": "Products", "range": "A1"},
                )
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        validation.validate_existence(workflow)  # should not raise

    def test_raises_when_sheet_is_missing(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="read_range",
                    params={"workbook": "wb", "sheet": "Nope", "range": "A1"},
                )
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        with pytest.raises(ValidationError) as exc_info:
            validation.validate_existence(workflow)
        assert "Nope" in exc_info.value.detail.message

    def test_a1_range_is_never_checked(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="read_range",
                    params={
                        "workbook": "wb",
                        "sheet": "Products",
                        "range": "ZZ9999:ZZ9999",
                    },
                )
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        validation.validate_existence(workflow)  # should not raise — A1-shaped, skipped

    def test_named_range_must_exist(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="read_range",
                    params={
                        "workbook": "wb",
                        "sheet": "Products",
                        "range": "NoSuchName",
                    },
                )
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        with pytest.raises(ValidationError) as exc_info:
            validation.validate_existence(workflow)
        assert "NoSuchName" in exc_info.value.detail.message

    def test_real_named_range_passes(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="read_range",
                    params={"workbook": "wb", "sheet": "Products", "range": "MyRange"},
                )
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        validation.validate_existence(workflow)  # should not raise

    def test_create_sheet_satisfies_a_later_reference(
        self, workbook_path: Path
    ) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="create_sheet",
                    params={"workbook": "wb", "name": "Archive"},
                ),
                Step(
                    id="s2",
                    action="write_cell",
                    params={
                        "workbook": "wb",
                        "sheet": "Archive",
                        "cell": "A1",
                        "value": 1,
                    },
                ),
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        validation.validate_existence(workflow)  # should not raise

    def test_write_to_sheet_not_yet_created_raises(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="write_cell",
                    params={
                        "workbook": "wb",
                        "sheet": "Archive",
                        "cell": "A1",
                        "value": 1,
                    },
                ),
                Step(
                    id="s2",
                    action="create_sheet",
                    params={"workbook": "wb", "name": "Archive"},
                ),
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        with pytest.raises(ValidationError):
            validation.validate_existence(workflow)

    def test_delete_sheet_then_reference_raises(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="delete_sheet",
                    params={"workbook": "wb", "sheet": "Products"},
                ),
                Step(
                    id="s2",
                    action="read_range",
                    params={"workbook": "wb", "sheet": "Products", "range": "A1"},
                ),
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        with pytest.raises(ValidationError):
            validation.validate_existence(workflow)

    def test_templated_sheet_is_skipped(self, workbook_path: Path) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="read_range",
                    params={
                        "workbook": "wb",
                        "sheet": "{{ steps.prior.output.sheet }}",
                        "range": "A1",
                    },
                )
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        validation.validate_existence(
            workflow
        )  # should not raise — can't know statically

    def test_nonexistent_workbook_file_is_skipped_entirely(self) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="read_range",
                    params={"workbook": "wb", "sheet": "Anything", "range": "A1"},
                )
            ],
            {
                "wb": WorkbookRef(
                    name="wb", file="C:/does/not/exist.xlsx", create_if_missing=True
                )
            },
        )
        validation.validate_existence(
            workflow
        )  # should not raise — nothing to check yet

    def test_copy_checks_both_source_and_target_sheets(
        self, workbook_path: Path
    ) -> None:
        workflow = _workflow(
            [
                Step(
                    id="s1",
                    action="copy",
                    params={
                        "source": {"workbook": "wb", "sheet": "Products"},
                        "target": {"workbook": "wb", "sheet": "Missing"},
                    },
                )
            ],
            {"wb": WorkbookRef(name="wb", file=str(workbook_path))},
        )
        with pytest.raises(ValidationError) as exc_info:
            validation.validate_existence(workflow)
        assert "Missing" in exc_info.value.detail.message
