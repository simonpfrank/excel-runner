"""Unit tests for backends.create_workbook (Spec sec 3, supports WorkbookRef.create_if_missing
via SessionManager, Spec sec 5.2)."""

from pathlib import Path

import openpyxl

from excel_runner import backends


class TestCreateWorkbook:
    def test_creates_a_blank_workbook(self, tmp_path: Path) -> None:
        path = tmp_path / "new.xlsx"
        backends.create_workbook(str(path))
        assert path.exists()
        workbook = openpyxl.load_workbook(path)
        assert workbook.sheetnames

    def test_creates_from_a_template(self, tmp_path: Path) -> None:
        template_path = tmp_path / "template.xlsx"
        template = openpyxl.Workbook()
        sheet = template.active
        assert sheet is not None
        sheet["A1"] = "Template content"
        template.save(template_path)

        new_path = tmp_path / "new.xlsx"
        backends.create_workbook(str(new_path), template_path=str(template_path))

        workbook = openpyxl.load_workbook(new_path)
        active = workbook.active
        assert active is not None
        assert active["A1"].value == "Template content"
