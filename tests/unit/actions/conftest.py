"""Shared fixtures for action tests: a real openpyxl fixture workbook wrapped in a
WorkbookSession, exactly what the (not-yet-built) runner will hand to an action function.
"""

from collections.abc import Iterator
from pathlib import Path

import openpyxl
import pytest

from excel_runner import backends
from excel_runner.backends import OwnedInstanceRegistry
from excel_runner.core import WorkbookSession


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "Region"
    sheet["B1"] = "Total"
    sheet["A2"] = "North"
    sheet["B2"] = 100
    workbook.save(path)
    return path


@pytest.fixture
def file_session(workbook_path: Path) -> Iterator[WorkbookSession]:
    handle = backends.open_workbook(str(workbook_path), mode="read_write")
    yield WorkbookSession(
        name="manip",
        backend="file",
        handle=handle,
        path=str(workbook_path),
        mode="read_write",
    )


@pytest.fixture
def richer_workbook_path(tmp_path: Path) -> Path:
    """A fixture with a real header row + data rows, for find_*/read_metadata tests."""
    path = tmp_path / "richer_fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "Notes"
    sheet["A2"] = "Region"
    sheet["B2"] = "Total"
    sheet["C2"] = "Status"
    sheet["A3"] = "North"
    sheet["B3"] = 100
    sheet["C3"] = "PASS"
    sheet["A4"] = "South"
    sheet["B4"] = 200
    sheet["C4"] = "FAIL"
    workbook.properties.title = "Q1 Report"
    workbook.properties.creator = "Simon"
    workbook.save(path)
    return path


@pytest.fixture
def richer_file_session(richer_workbook_path: Path) -> Iterator[WorkbookSession]:
    handle = backends.open_workbook(str(richer_workbook_path), mode="read_write")
    yield WorkbookSession(
        name="manip",
        backend="file",
        handle=handle,
        path=str(richer_workbook_path),
        mode="read_write",
    )


@pytest.fixture
def multi_sheet_workbook_path(tmp_path: Path) -> Path:
    """Three sheets, two sharing an "A&H" prefix — the exact shape step 15's multi-sheet
    capture needs (list/all/matching sheet selection for `read_range`)."""
    path = tmp_path / "multi_sheet_fixture.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    assert first is not None
    first.title = "A&H North"
    first["A1"] = "north-value"
    second = workbook.create_sheet("A&H South")
    second["A1"] = "south-value"
    third = workbook.create_sheet("Other")
    third["A1"] = "other-value"
    workbook.save(path)
    return path


@pytest.fixture
def multi_sheet_file_session(
    multi_sheet_workbook_path: Path,
) -> Iterator[WorkbookSession]:
    handle = backends.open_workbook(str(multi_sheet_workbook_path), mode="read_write")
    yield WorkbookSession(
        name="manip",
        backend="file",
        handle=handle,
        path=str(multi_sheet_workbook_path),
        mode="read_write",
    )


@pytest.fixture
def formula_workbook_path(tmp_path: Path) -> Path:
    """B1 = A1*2, with a *real* cached value from a real Excel recalculation — not the stale
    cache openpyxl alone can produce (it never computes formulas), so this genuinely
    distinguishes "read the cached value" from "read the formula text" (data_only=True vs
    False), the exact bug this fixture exists to catch."""
    path = tmp_path / "formula_fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = 10
    sheet["B1"] = "=A1*2"
    workbook.save(path)

    registry = OwnedInstanceRegistry()
    app = registry.spawn()
    try:
        book = backends.xlw_open_workbook(app, str(path), mode="read_write")
        backends.com_calculate_workbook(book)
        backends.com_wait_until_calculation_done(app)
        backends.xlw_save_workbook(book)
        backends.xlw_close_workbook(book)
    finally:
        registry.close_owned()
    return path


@pytest.fixture
def formula_file_session(formula_workbook_path: Path) -> Iterator[WorkbookSession]:
    handle = backends.open_workbook(str(formula_workbook_path), mode="read_write")
    yield WorkbookSession(
        name="manip",
        backend="file",
        handle=handle,
        path=str(formula_workbook_path),
        mode="read_write",
    )


@pytest.fixture
def named_range_workbook_path(tmp_path: Path) -> Path:
    """A workbook with a workbook-level defined name ("SalesTotal") pointing at Summary!B2."""
    path = tmp_path / "named_range_fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "Region"
    sheet["B1"] = "Total"
    sheet["A2"] = "North"
    sheet["B2"] = 100
    workbook.defined_names["SalesTotal"] = openpyxl.workbook.defined_name.DefinedName(
        "SalesTotal", attr_text="Summary!$B$2"
    )
    workbook.save(path)
    return path


@pytest.fixture
def named_range_file_session(
    named_range_workbook_path: Path,
) -> Iterator[WorkbookSession]:
    handle = backends.open_workbook(str(named_range_workbook_path), mode="read_write")
    yield WorkbookSession(
        name="manip",
        backend="file",
        handle=handle,
        path=str(named_range_workbook_path),
        mode="read_write",
    )


@pytest.fixture
def richer_named_range_workbook_path(tmp_path: Path) -> Path:
    """The richer_workbook_path shape, plus two defined names: "HeaderSearchArea" (a multi-row
    range, for find_headers_row) and "NotesCell" (a single cell, for read_metadata cells).
    """
    path = tmp_path / "richer_named_range_fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = "Notes"
    sheet["A2"] = "Region"
    sheet["B2"] = "Total"
    sheet["C2"] = "Status"
    sheet["A3"] = "North"
    sheet["B3"] = 100
    sheet["C3"] = "PASS"
    sheet["A4"] = "South"
    sheet["B4"] = 200
    sheet["C4"] = "FAIL"
    workbook.defined_names["HeaderSearchArea"] = (
        openpyxl.workbook.defined_name.DefinedName(
            "HeaderSearchArea", attr_text="Summary!$A$1:$C$4"
        )
    )
    workbook.defined_names["NotesCell"] = openpyxl.workbook.defined_name.DefinedName(
        "NotesCell", attr_text="Summary!$A$1"
    )
    workbook.save(path)
    return path


@pytest.fixture
def richer_named_range_file_session(
    richer_named_range_workbook_path: Path,
) -> Iterator[WorkbookSession]:
    handle = backends.open_workbook(
        str(richer_named_range_workbook_path), mode="read_write"
    )
    yield WorkbookSession(
        name="manip",
        backend="file",
        handle=handle,
        path=str(richer_named_range_workbook_path),
        mode="read_write",
    )
