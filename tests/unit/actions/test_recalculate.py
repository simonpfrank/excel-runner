"""Unit tests for the `recalculate` action (Spec sec 7 catalog). Real xlwings against a real,
locally-spawned Excel instance -- no mocks (project convention, matches
test_owned_instance_registry.py / test_backends_xlw.py)."""

from collections.abc import Iterator
from pathlib import Path

import openpyxl
import pytest

from excel_runner.actions import recalculate as recalculate_action
from excel_runner.backends import OwnedInstanceRegistry, xlw_close_workbook, xlw_open_workbook
from excel_runner.core import ACTION_CAPABILITIES, ActionExecutionError, WorkbookSession
from tests.unit.conftest import requires_excel, requires_working_xlwings_save


@pytest.fixture
def formula_workbook_path(tmp_path: Path) -> Path:
    """A1 feeds B1 (=A1*2); openpyxl can't compute the formula, so B1's cached value is
    deliberately stale until a real recalculation runs."""
    path = tmp_path / "fixture.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Summary"
    sheet["A1"] = 10
    sheet["B1"] = "=A1*2"
    other = workbook.create_sheet("Other")
    other["A1"] = 3
    other["B1"] = "=A1*10"
    workbook.save(path)
    return path


@pytest.fixture
def xlw_session(formula_workbook_path: Path) -> Iterator[WorkbookSession]:
    registry = OwnedInstanceRegistry()
    app = registry.spawn()
    book = xlw_open_workbook(app, str(formula_workbook_path), mode="read_write")
    try:
        yield WorkbookSession(
            name="manip",
            backend="xlw",
            handle=book,
            path=str(formula_workbook_path),
            mode="read_write",
        )
    finally:
        xlw_close_workbook(book)
        registry.close_owned()


class TestRecalculateAction:
    def test_registers_as_a_com_action(self) -> None:
        assert ACTION_CAPABILITIES["recalculate"] == "com"


@requires_excel
@requires_working_xlwings_save
class TestRecalculateBehavior:
    def test_default_scope_and_mode_recalculates_the_workbook(
        self, xlw_session: WorkbookSession
    ) -> None:
        xlw_session.handle.sheets["Summary"]["A1"].value = 21
        xlw_session.handle.sheets["Other"]["A1"].value = 4

        result = recalculate_action(session=xlw_session)

        assert result.status == "success"
        assert result.output == {"scope": "workbook", "mode": "normal"}
        assert xlw_session.handle.sheets["Summary"]["B1"].value == 42
        # scope: workbook recalculates every sheet in *this* workbook too.
        assert xlw_session.handle.sheets["Other"]["B1"].value == 40

    def test_scope_sheet_recalculates_only_the_named_sheet(
        self, xlw_session: WorkbookSession
    ) -> None:
        xlw_session.handle.sheets["Summary"]["A1"].value = 5
        xlw_session.handle.sheets["Other"]["A1"].value = 100

        result = recalculate_action(session=xlw_session, scope="sheet", sheet="Summary")

        assert result.output["sheet"] == "Summary"
        assert "warning" not in result.output
        assert xlw_session.handle.sheets["Summary"]["B1"].value == 10

    def test_scope_sheet_without_a_name_falls_back_to_the_active_sheet_with_a_warning(
        self, xlw_session: WorkbookSession
    ) -> None:
        xlw_session.handle.sheets["Summary"].activate()
        xlw_session.handle.sheets["Summary"]["A1"].value = 6

        result = recalculate_action(session=xlw_session, scope="sheet")

        assert result.output["sheet"] == "Summary"
        assert "Summary" in result.output["warning"]
        assert xlw_session.handle.sheets["Summary"]["B1"].value == 12

    def test_scope_all_mode_full_recalculates(self, xlw_session: WorkbookSession) -> None:
        xlw_session.handle.sheets["Summary"]["A1"].value = 7

        result = recalculate_action(session=xlw_session, scope="all", mode="full")

        assert result.output == {"scope": "all", "mode": "full"}
        assert xlw_session.handle.sheets["Summary"]["B1"].value == 14

    def test_scope_all_mode_full_rebuild_recalculates(self, xlw_session: WorkbookSession) -> None:
        xlw_session.handle.sheets["Summary"]["A1"].value = 9

        recalculate_action(session=xlw_session, scope="all", mode="full_rebuild")

        assert xlw_session.handle.sheets["Summary"]["B1"].value == 18

    def test_always_saves_before_returning(self, xlw_session: WorkbookSession) -> None:
        xlw_session.handle.sheets["Summary"]["A1"].value = 11
        xlw_session.dirty = True

        recalculate_action(session=xlw_session)

        assert xlw_session.dirty is False
        reopened = openpyxl.load_workbook(xlw_session.path, data_only=True)
        assert reopened["Summary"]["B1"].value == 22

    def test_sheet_param_with_scope_workbook_is_rejected(
        self, xlw_session: WorkbookSession
    ) -> None:
        with pytest.raises(ActionExecutionError) as exc_info:
            recalculate_action(session=xlw_session, scope="workbook", sheet="Summary")
        assert "sheet" in exc_info.value.detail.message.lower()

    def test_mode_full_with_scope_workbook_is_rejected(
        self, xlw_session: WorkbookSession
    ) -> None:
        with pytest.raises(ActionExecutionError) as exc_info:
            recalculate_action(session=xlw_session, scope="workbook", mode="full")
        assert "application-wide" in exc_info.value.detail.message.lower()

    def test_mode_full_rebuild_with_scope_sheet_is_rejected(
        self, xlw_session: WorkbookSession
    ) -> None:
        with pytest.raises(ActionExecutionError):
            recalculate_action(session=xlw_session, scope="sheet", sheet="Summary", mode="full_rebuild")
